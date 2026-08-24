#!/usr/bin/env python3
from __future__ import annotations

import argparse
import io
import json
import math
import re
import sys
import unicodedata
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UA = 'CEPOES-data-bot/1.0 (+https://cepoes.org/)'
TIMEOUT = 60

M01_PAGE = 'https://www.estadisticaciudad.gob.ar/eyc/banco-datos/distribucion-porcentual-de-la-poblacion-por-lugar-de-nacimiento-segun-comuna-ciudad-de-buenos-aires-anos-2006-20082016/'
INDICATORS = {
    'countries': 'https://www.estadisticaciudad.gob.ar/si/demog/principal-indicador?indicador=b28',
    'activity': 'https://www.estadisticaciudad.gob.ar/si/demog/principal-indicador?indicador=b27a',
    'schooling': 'https://www.estadisticaciudad.gob.ar/si/demog/principal-indicador?indicador=b29b',
    'poverty': 'https://www.estadisticaciudad.gob.ar/si/demog/principal-indicador?indicador=b211b',
    'recent': 'https://www.estadisticaciudad.gob.ar/si/demog/principal-indicador?indicador=b25',
}


def norm(value) -> str:
    s = '' if value is None else str(value)
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s).strip().lower()


def number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)
    s = str(value).strip().replace('\xa0', ' ')
    m = re.search(r'-?\d+(?:[\.,]\d+)?', s)
    if not m:
        return None
    token = m.group(0)
    if ',' in token and '.' in token:
        token = token.replace('.', '').replace(',', '.')
    elif ',' in token:
        token = token.replace(',', '.')
    try:
        return float(token)
    except ValueError:
        return None


def get(session: requests.Session, url: str, *, binary=False):
    last = None
    for i in range(3):
        try:
            r = session.get(url, timeout=TIMEOUT, headers={'User-Agent': UA})
            r.raise_for_status()
            return r.content if binary else r.text
        except Exception as e:
            last = e
    raise RuntimeError(f'No se pudo descargar {url}: {last}')


def parse_years_and_rows(html: str):
    soup = BeautifulSoup(html, 'html.parser')
    candidates = []
    for table in soup.find_all('table'):
        rows = []
        for tr in table.find_all('tr'):
            cells = [re.sub(r'\s+', ' ', c.get_text(' ', strip=True)).strip() for c in tr.find_all(['th', 'td'])]
            if cells:
                rows.append(cells)
        if rows:
            candidates.append(rows)
    if candidates:
        rows = max(candidates, key=lambda rs: sum(len(r) for r in rs))
        yidx = None
        years = []
        for i, row in enumerate(rows):
            ys = [int(x) for x in row if re.fullmatch(r'(?:19|20)\d{2}', x)]
            if len(ys) >= 2:
                yidx, years = i, ys
                break
        if yidx is not None:
            out = {}
            for row in rows[yidx+1:]:
                label = next((x for x in row if x and not re.fullmatch(r'-?\d+(?:[\.,]\d+)?(?:\s*[a-z])?', x, re.I)), None)
                nums = [number(x) for x in row]
                nums = [x for x in nums if x is not None]
                if label and len(nums) >= len(years):
                    out[label] = nums[-len(years):]
            if out:
                return years, out
    lines = [re.sub(r'\s+', ' ', x).strip() for x in soup.get_text('\n').splitlines() if re.sub(r'\s+', ' ', x).strip()]
    for i, line in enumerate(lines):
        ys = [int(x) for x in re.findall(r'(?<!\d)((?:19|20)\d{2})(?!\d)', line)]
        if len(ys) >= 2:
            years = ys
            out = {}
            j = i + 1
            while j < len(lines):
                label = lines[j]
                if 'unidad de medida' in norm(label) or 'fuente:' in norm(label):
                    break
                vals=[]; k=j+1
                while k < len(lines) and len(vals) < len(years):
                    v=number(lines[k])
                    if v is None: break
                    vals.append(v); k+=1
                if len(vals)==len(years):
                    out[label]=vals; j=k
                else:
                    nums=[number(x) for x in re.findall(r'-?\d+(?:[\.,]\d+)?', label)]
                    nums=[x for x in nums if x is not None]
                    cleanlabel=re.sub(r'(?:\s+-?\d+(?:[\.,]\d+)?(?:\s*[a-z])?)+\s*$', '', label, flags=re.I).strip()
                    if cleanlabel and len(nums)>=len(years): out[cleanlabel]=nums[-len(years):]
                    j+=1
            if out:
                return years,out
    raise RuntimeError('No se pudo interpretar la tabla del indicador')


def pick_row(rows, *aliases):
    wanted = [norm(a) for a in aliases]
    for label, vals in rows.items():
        nl = norm(label)
        if any(nl == a or a in nl for a in wanted):
            return vals
    raise KeyError(f'No se encontró fila {aliases}; disponibles={list(rows)[:30]}')


def parse_indicator(session, url):
    html = get(session, url)
    return parse_years_and_rows(html)


def find_xlsx(session, page_url):
    html = get(session, page_url)
    soup = BeautifulSoup(html, 'html.parser')
    for a in soup.find_all('a', href=True):
        href=a['href']
        if '.xlsx' in href.lower():
            return requests.compat.urljoin(page_url, href)
    m=re.search(r'https?://[^\"\']+\.xlsx(?:\?[^\"\']*)?', html, re.I)
    if m: return m.group(0)
    raise RuntimeError('No se encontró el XLSX oficial de lugar de nacimiento por comuna')


def rows_from_workbook(content: bytes):
    wb=load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    all_rows=[]
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals=list(row)
            if any(v not in (None,'') for v in vals):
                all_rows.append(vals)
    return all_rows


def parse_commune_eah(content: bytes):
    rows=rows_from_workbook(content)
    aliases={
      'nacida_caba_pct':['ciudad de buenos aires','en esta ciudad','caba'],
      'prov_ba_pct':['provincia de buenos aires','pcia. de buenos aires','prov. de buenos aires','prov. buenos aires','prov buenos aires'],
      'otra_provincia_pct':['otra provincia','otras provincias'],
      'pais_limitrofe_pct':['pais limitrofe','país limítrofe','paises limitrofes'],
      'pais_no_limitrofe_pct':['pais no limitrofe','otro pais','otros paises','resto de paises'],
    }
    best=None
    for ri,row in enumerate(rows):
        mapped={}
        for ci,v in enumerate(row):
            nv=norm(v)
            for key,names in aliases.items():
                if any(norm(name) in nv for name in names):
                    mapped.setdefault(key,ci)
        if len(mapped)>=4 and (best is None or len(mapped)>len(best[1])):
            best=(ri,mapped,row)
    if best is None:
        sample=[' | '.join(str(v) for v in r if v not in (None,''))[:300] for r in rows[:80]]
        raise RuntimeError('M01: no se encontró encabezado de categorías. Primeras filas:\n'+'\n'.join(sample))
    ri,cols,header=best
    if len(cols)<5:
        raise RuntimeError(f'M01: encabezado incompleto: {cols}; fila={header}')

    data=rows[ri+1:]
    years_found=[]
    for r in rows:
        for v in r:
            if isinstance(v,(int,float)) and 2000 <= float(v) <= 2100:
                years_found.append(int(v))
            else:
                sv=str(v).strip() if v is not None else ''
                if re.fullmatch(r'(?:19|20)\d{2}', sv):
                    years_found.append(int(sv))
    if not years_found:
        raise RuntimeError('M01: no se encontró ningún año en el XLSX oficial')
    latest_year=max(years_found)
    results={}
    current_year=None
    for row in data:
        year_cells=[int(v) for v in row[:6] if isinstance(v,(int,float)) and 2000<=v<=2100]
        if year_cells: current_year=year_cells[0]
        if current_year is None:
            for v in row[:6]:
                if re.fullmatch(r'(?:19|20)\d{2}', str(v).strip() if v is not None else ''):
                    current_year=int(v); break
        comuna=None
        for v in row[:8]:
            if isinstance(v,(int,float)) and float(v).is_integer() and 1<=int(v)<=15:
                comuna=str(int(v)); break
            sv=norm(v)
            m=re.fullmatch(r'(?:comuna\s*)?(1[0-5]|[1-9])', sv)
            if m: comuna=m.group(1); break
        if comuna is None or current_year!=latest_year:
            continue
        vals={}
        for key,ci in cols.items():
            if ci < len(row): vals[key]=number(row[ci])
        if all(vals.get(k) is not None for k in aliases):
            vals={k:round(float(vals[k]),2) for k in aliases}
            vals['migracion_interna_pct']=round(vals['prov_ba_pct']+vals['otra_provincia_pct'],2)
            vals['migracion_internacional_pct']=round(vals['pais_limitrofe_pct']+vals['pais_no_limitrofe_pct'],2)
            results[comuna]=vals
    if len(results)!=15:
        candidates=[]
        for row in data:
            comuna=None
            for v in row[:8]:
                if isinstance(v,(int,float)) and float(v).is_integer() and 1<=int(v)<=15:
                    comuna=str(int(v)); break
                m=re.fullmatch(r'(?:comuna\s*)?(1[0-5]|[1-9])', norm(v))
                if m: comuna=m.group(1); break
            if not comuna: continue
            vals={k:number(row[ci]) if ci<len(row) else None for k,ci in cols.items()}
            if all(vals.get(k) is not None and 0<=vals[k]<=100 for k in aliases):
                candidates.append((comuna,{k:round(float(vals[k]),2) for k in aliases}))
        runs=[]; cur=[]; seen=set()
        for item in candidates:
            c=item[0]
            if c in seen:
                if len(seen)>=12: runs.append(cur)
                cur=[];seen=set()
            cur.append(item);seen.add(c)
            if len(seen)==15:
                runs.append(cur);cur=[];seen=set()
        if runs:
            run=runs[-1]
            results={}
            for c,vals in run:
                vals['migracion_interna_pct']=round(vals['prov_ba_pct']+vals['otra_provincia_pct'],2)
                vals['migracion_internacional_pct']=round(vals['pais_limitrofe_pct']+vals['pais_no_limitrofe_pct'],2)
                results[c]=vals
    if len(results)!=15:
        raise RuntimeError(f'M01: se esperaban 15 comunas para {latest_year}, se obtuvieron {len(results)}: {sorted(results)}')
    total={}
    for row in reversed(data):
        label=' '.join(norm(v) for v in row[:8] if v not in (None,''))
        if 'total' not in label and 'ciudad' not in label: continue
        vals={k:number(row[ci]) if ci<len(row) else None for k,ci in cols.items()}
        if all(vals.get(k) is not None and 0<=vals[k]<=100 for k in aliases):
            total={k:round(float(vals[k]),2) for k in aliases}
            total['migracion_interna_pct']=round(total['prov_ba_pct']+total['otra_provincia_pct'],2)
            total['migracion_internacional_pct']=round(total['pais_limitrofe_pct']+total['pais_no_limitrofe_pct'],2)
            break
    return latest_year,results,total


def current_or_error(years, rows, aliases):
    year=max(years)
    idx=years.index(year)
    vals=pick_row(rows,*aliases)
    return year, float(vals[idx])


def build(previous: dict, session: requests.Session):
    out=deepcopy(previous)
    out['schema']='cepoes-migraciones-v1'
    out['status']='VALIDADO'
    out['updated_at']=datetime.now(timezone.utc).isoformat().replace('+00:00','Z')
    out['generated']=datetime.now(timezone.utc).date().isoformat()
    out.setdefault('latest',{})

    xlsx_url=find_xlsx(session,M01_PAGE)
    xlsx=get(session,xlsx_url,binary=True)
    year,communes,total=parse_commune_eah(xlsx)
    prev_year=int(previous.get('headline',{}).get('eah',{}).get('year',0) or 0)
    if year < prev_year: raise RuntimeError(f'M01 retrocedió de {prev_year} a {year}')
    for c,vals in communes.items():
        out['communes'][c]['eah']={'year':year,**vals}
    if total:
        out['headline']['eah']={'year':year,**total}
    else:
        if year!=prev_year:
            raise RuntimeError('M01 publicó un año nuevo pero no se pudo identificar el total Ciudad')
        out['headline']['eah']['year']=year
    out['latest']['place_birth_commune_eah']=year

    years,rows=parse_indicator(session,INDICATORS['countries'])
    cy=max(years); ci=years.index(cy)
    base=min(years); bi=years.index(base)
    country_names=['Bolivia','Brasil','Colombia','Chile','Paraguay','Perú','Uruguay','Venezuela','España','Italia','Otros países']
    crow=[]
    for name in country_names:
        vals=pick_row(rows,name)
        cur=float(vals[ci]); old=float(vals[bi])
        crow.append({'pais':name,'pct_2024':round(cur,2),'pct_2015':round(old,2),'cambio_pp':round(cur-old,2)})
    out['countries']={'year':cy,'base_year':base,'rows':crow}
    out['latest']['countries']=cy

    years,rows=parse_indicator(session,INDICATORS['recent'])
    order=sorted(range(len(years)), key=lambda i: years[i])
    sy=[years[i] for i in order]
    def seq(*names):
        vals=pick_row(rows,*names); return [round(float(vals[i]),2) for i in order]
    out['recent_migration']={
      'definition':'Población de 5 años y más que residía fuera de CABA cinco años antes del censo.',
      'years':sy,
      'prov_ba':seq('Provincia de Buenos Aires'),
      'otra_provincia':seq('Otra provincia'),
      'exterior':seq('En el extranjero','Exterior'),
    }
    out['latest']['recent_migration']=max(sy)

    years,rows=parse_indicator(session,INDICATORS['activity']); y=max(years); i=years.index(y)
    out['socioeconomic']['activity']={'year':y,'values':{
      'total':round(float(pick_row(rows,'TOTAL')[i]),2),
      'caba':round(float(pick_row(rows,'Ciudad de Buenos Aires')[i]),2),
      'resto_pais':round(float(pick_row(rows,'Resto del país')[i]),2),
      'exterior':round(float(pick_row(rows,'Exterior')[i]),2),
    }}; out['latest']['activity']=y

    years,rows=parse_indicator(session,INDICATORS['schooling']); y=max(years); i=years.index(y)
    out['socioeconomic']['schooling']={'year':y,'unit':'años promedio','values':{
      'total':round(float(pick_row(rows,'TOTAL')[i]),2),
      'caba':round(float(pick_row(rows,'Ciudad de Buenos Aires')[i]),2),
      'resto_pais':round(float(pick_row(rows,'Resto del país')[i]),2),
      'exterior':round(float(pick_row(rows,'Exterior')[i]),2),
    }}; out['latest']['schooling']=y

    years,rows=parse_indicator(session,INDICATORS['poverty']); y=max(years); i=years.index(y)
    out['socioeconomic']['poverty_multidimensional']={'year':y,'values':{
      'total':round(float(pick_row(rows,'TOTAL')[i]),2),
      'caba':round(float(pick_row(rows,'En esta ciudad','Ciudad de Buenos Aires')[i]),2),
      'prov_ba':round(float(pick_row(rows,'En la Provincia de Buenos Aires','Provincia de Buenos Aires')[i]),2),
      'otra_provincia':round(float(pick_row(rows,'En otra provincia','Otra provincia')[i]),2),
      'pais_limitrofe':round(float(pick_row(rows,'En país limítrofe','País limítrofe')[i]),2),
      'otro_pais':round(float(pick_row(rows,'En otro país','Otro país')[i]),2),
    }}; out['latest']['poverty_multidimensional']=y

    out['automation']={
      'mode':'daily-source-check',
      'publication_rule':'Sólo se publica una actualización cuando las fuentes oficiales superan controles de estructura, cobertura y consistencia.',
      'fallback':'Ante un cambio de formato o error de descarga se conserva el último dato validado.',
      'm01_download':xlsx_url,
      'checked_at':out['updated_at'],
    }
    return out


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--previous',default='deploy/site-overlay/assets/data/migraciones.json')
    ap.add_argument('--output',default='datos/migraciones/migraciones.json')
    ap.add_argument('--overlay',default='deploy/site-overlay/assets/data/migraciones.json')
    args=ap.parse_args()
    previous_path=Path(args.previous)
    if not previous_path.exists():
        raise SystemExit(f'Falta baseline: {previous_path}')
    previous=json.loads(previous_path.read_text(encoding='utf-8'))
    session=requests.Session()
    out=build(previous,session)
    text=json.dumps(out,ensure_ascii=False,indent=2)+'\n'
    for dest in [Path(args.output),Path(args.overlay)]:
        dest.parent.mkdir(parents=True,exist_ok=True); dest.write_text(text,encoding='utf-8')
    print('Migraciones actualizado:', out['latest'])

if __name__=='__main__':
    main()
