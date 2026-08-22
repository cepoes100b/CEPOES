from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook

import oferta_extra_2_fix  # noqa: F401 — aplica altas/correcciones verificadas al catálogo
from fuentes_territorio import DATASETS_TERRITORIO
from oferta_extra_2 import EXTRA_DATASETS

CPA_RE = re.compile(r"^C\d{4}[A-Z]{3}$", re.I)
CPA_HEADERS = {
    "cpa", "codigo_postal_argentino", "cod_postal_argentino", "codigo_postal", "cod_postal",
    "codigo_postal_arg", "codigopostalargentino",
}
BARRIO_HEADERS = {"barrio", "barrios", "nombre_barrio", "barrio_nombre"}
COMUNA_HEADERS = {"comuna", "nro_comuna", "numero_comuna", "comuna_nro", "comuna_numero"}


def norm_text(value: object) -> str:
    s = str(value or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def norm_barrio(value: object) -> str:
    return norm_text(value).replace("_de_", "_").replace("_del_", "_")


def norm_cpa(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().upper())


def parse_comuna(value: object) -> int | None:
    if value is None:
        return None
    m = re.search(r"\b(1[0-5]|[1-9])\b", str(value))
    return int(m.group(1)) if m else None


def cargar_barrios(path: Path) -> tuple[dict[str, tuple[str, int]], set[str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    barrios = data.get("barrios") or {}
    out: dict[str, tuple[str, int]] = {}
    nombres: set[str] = set()
    for item in barrios.values():
        nombre = str(item.get("nombre") or "").strip()
        comuna = int(item.get("comuna"))
        if not nombre or not 1 <= comuna <= 15:
            continue
        key = norm_barrio(nombre)
        out[key] = (nombre, comuna)
        nombres.add(nombre)
    if len(nombres) != 48:
        raise ValueError(f"Se esperaban 48 barrios canónicos; obtenidos {len(nombres)}")

    # Alias puramente nominales observados habitualmente en datos GCBA.
    aliases = {
        "villa_gral_mitre": "villa_general_mitre",
        "villa_gral_mitre_": "villa_general_mitre",
        "paternal": "la_paternal",
        "nueva_pompeya": "nueva_pompeya",
        "parque_avellaneda": "parque_avellaneda",
        "v_sarsfield": "velez_sarsfield",
    }
    for alias, canonical in aliases.items():
        if canonical in out:
            out[alias] = out[canonical]
    return out, nombres


def iter_csv(path: Path) -> Iterator[dict[str, object]]:
    raw = path.read_bytes()[:65536]
    text = None
    encoding = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None or encoding is None:
        raise ValueError("CSV sin codificación reconocible")
    try:
        dialect = csv.Sniffer().sniff(text, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if text.count(";") > text.count(",") else ","
    with path.open("r", encoding=encoding, errors="replace", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        for row in reader:
            yield dict(row)


def iter_xlsx(path: Path) -> Iterator[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return
        headers = [str(v or "").strip() for v in header]
        for values in rows:
            yield {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
    finally:
        wb.close()


def iter_rows(path: Path) -> Iterator[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        yield from iter_xlsx(path)
    else:
        return


def pick_header(headers: Iterable[object], candidates: set[str]) -> str | None:
    mapping = {norm_text(h): str(h) for h in headers if str(h or "").strip()}
    for candidate in candidates:
        if candidate in mapping:
            return mapping[candidate]
    # Tolerancia a prefijos/sufijos descriptivos, sin aceptar cualquier campo que contenga CP.
    for normalized, original in mapping.items():
        if candidates is CPA_HEADERS and ("codigo_postal_argentino" in normalized or normalized == "cpa"):
            return original
        if candidates is BARRIO_HEADERS and (normalized == "barrio" or normalized.endswith("_barrio")):
            return original
        if candidates is COMUNA_HEADERS and normalized.startswith("comuna"):
            return original
    return None


def fuentes_configuradas() -> dict[str, dict]:
    merged: dict[str, dict] = {}
    for family, source in (("territorio", DATASETS_TERRITORIO), ("oferta_extra", EXTRA_DATASETS)):
        for key, cfg in source.items():
            entry = dict(cfg)
            entry["familia"] = family
            merged[key] = entry
    return merged


def construir(badata: Path, territorio: Path) -> tuple[list[dict], dict]:
    canonical, nombres_canonicos = cargar_barrios(territorio)
    configs = fuentes_configuradas()

    observations: dict[str, Counter[tuple[str, int]]] = defaultdict(Counter)
    provenance: dict[tuple[str, str, int], set[str]] = defaultdict(set)
    source_stats: list[dict] = []
    unknown_barrio = Counter()
    total_rows = 0

    for key, cfg in sorted(configs.items()):
        path = badata / cfg["filename"]
        stat = {
            "fuente": key,
            "dataset": cfg.get("dataset"),
            "archivo": cfg["filename"],
            "existe": path.exists(),
            "filas": 0,
            "tiene_cpa": False,
            "tiene_barrio": False,
            "tiene_comuna": False,
            "observaciones_validas": 0,
            "cpa_distintos_validos": 0,
            "barrios_no_canonicos": 0,
            "comuna_inconsistente": 0,
        }
        if not path.exists():
            source_stats.append(stat)
            continue
        try:
            rows = iter_rows(path)
            first = next(rows, None)
            if first is None:
                source_stats.append(stat)
                continue
            headers = list(first.keys())
            cpa_h = pick_header(headers, CPA_HEADERS)
            barrio_h = pick_header(headers, BARRIO_HEADERS)
            comuna_h = pick_header(headers, COMUNA_HEADERS)
            stat["tiene_cpa"] = bool(cpa_h)
            stat["tiene_barrio"] = bool(barrio_h)
            stat["tiene_comuna"] = bool(comuna_h)
            if not cpa_h or not barrio_h:
                source_stats.append(stat)
                continue

            seen_source: set[str] = set()
            for row in (first, *rows):
                stat["filas"] += 1
                total_rows += 1
                cpa = norm_cpa(row.get(cpa_h))
                if not CPA_RE.fullmatch(cpa):
                    continue
                barrio_raw = str(row.get(barrio_h) or "").strip()
                canon = canonical.get(norm_barrio(barrio_raw))
                if not canon:
                    stat["barrios_no_canonicos"] += 1
                    if barrio_raw:
                        unknown_barrio[barrio_raw] += 1
                    continue
                barrio, comuna = canon
                if comuna_h:
                    fuente_comuna = parse_comuna(row.get(comuna_h))
                    if fuente_comuna is not None and fuente_comuna != comuna:
                        stat["comuna_inconsistente"] += 1
                        continue
                observations[cpa][(barrio, comuna)] += 1
                provenance[(cpa, barrio, comuna)].add(key)
                stat["observaciones_validas"] += 1
                seen_source.add(cpa)
            stat["cpa_distintos_validos"] = len(seen_source)
        except Exception as exc:
            stat["error"] = f"{type(exc).__name__}: {exc}"
        source_stats.append(stat)

    rows_out: list[dict] = []
    conflicts: list[dict] = []
    for cpa, options in sorted(observations.items()):
        if len(options) != 1:
            conflicts.append({
                "cpa": cpa,
                "asignaciones": [
                    {"barrio": b, "comuna": c, "observaciones": n}
                    for (b, c), n in options.most_common()
                ],
            })
            continue
        (barrio, comuna), count = next(iter(options.items()))
        sources = sorted(provenance[(cpa, barrio, comuna)])
        rows_out.append({
            "cpa": cpa,
            "barrio": barrio,
            "comuna": comuna,
            "fuentes_n": len(sources),
            "observaciones": count,
            "fuentes": "|".join(sources),
        })

    covered_barrios = sorted({r["barrio"] for r in rows_out})
    state = {
        "schema": 1,
        "producto": "cruce_cpa_barrio_gcba_observado",
        "criterio": "Sólo CPA completos observados en fuentes oficiales GCBA con barrio canónico; conflictos excluidos",
        "fuentes_configuradas": len(configs),
        "fuentes_con_archivo": sum(1 for s in source_stats if s["existe"]),
        "fuentes_con_cpa_y_barrio": sum(1 for s in source_stats if s["tiene_cpa"] and s["tiene_barrio"]),
        "filas_revisadas_en_fuentes_utiles": total_rows,
        "cpa_observados": len(observations),
        "cpa_utilizables": len(rows_out),
        "cpa_conflictivos": len(conflicts),
        "barrios_cubiertos": len(covered_barrios),
        "barrios_canonicos_total": len(nombres_canonicos),
        "barrios_sin_cpa_observado": sorted(nombres_canonicos - set(covered_barrios)),
        "fuentes": source_stats,
        "conflictos": conflicts,
        "barrios_no_canonicos_mas_frecuentes": [
            {"valor": k, "observaciones": v} for k, v in unknown_barrio.most_common(25)
        ],
    }
    return rows_out, state


def escribir_csv(path: Path, rows: list[dict]) -> None:
    fields = ["cpa", "barrio", "comuna", "fuentes_n", "observaciones", "fuentes"]
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    ap = argparse.ArgumentParser(description="Construye un cruce CPA→barrio/comuna a partir de observaciones oficiales GCBA")
    ap.add_argument("--badata", type=Path, default=Path("badata"))
    ap.add_argument("--territorio", type=Path, default=Path("territorio.json"))
    ap.add_argument("--salida", type=Path, default=Path("cpa_territorio.csv"))
    ap.add_argument("--estado", type=Path, default=Path("estado_cpa_territorio.json"))
    args = ap.parse_args()

    rows, state = construir(args.badata, args.territorio)
    escribir_csv(args.salida, rows)
    args.estado.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"CPA territorial · {state['cpa_utilizables']} utilizables · {state['cpa_conflictivos']} conflictivos · "
        f"{state['barrios_cubiertos']}/48 barrios · {state['fuentes_con_cpa_y_barrio']} fuentes útiles"
    )
    if state["barrios_sin_cpa_observado"]:
        print("Barrios sin CPA observado:", ", ".join(state["barrios_sin_cpa_observado"]))
    if state["barrios_no_canonicos_mas_frecuentes"]:
        vals = state["barrios_no_canonicos_mas_frecuentes"][:10]
        print("Valores de barrio no canónicos más frecuentes:", vals)


if __name__ == "__main__":
    main()
