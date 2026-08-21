"""Descarga fuentes territoriales desde BA Data (CKAN).

Cada recurso se resuelve por el nombre dentro de su dataset. Si una descarga
falla, conserva la copia commiteada anterior. El generador puede trabajar sin
conexión con esas copias.
"""
from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

from fuentes_territorio import BA_DATA_API, DATASETS_TERRITORIO

BASE = Path(__file__).resolve().parent
DIR = BASE / "badata"
DIR.mkdir(exist_ok=True)
ESTADO = BASE / "estado_territorio.json"

try:
    sys.stdout.reconfigure(line_buffering=True)
except AttributeError:
    pass

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "CEPOES-data-pipeline/1.0 (+https://cepoes.org)",
    "Accept": "application/json,text/plain,*/*",
})
TIMEOUT = (10, 90)
RETRIES = 3


def get(url: str, **kwargs):
    last = None
    for i in range(RETRIES):
        try:
            r = SESSION.get(url, timeout=TIMEOUT, **kwargs)
            if r.status_code == 200:
                return r
            last = RuntimeError(f"HTTP {r.status_code}")
        except requests.RequestException as e:
            last = e
        time.sleep(2 * (i + 1))
    raise RuntimeError(str(last or "sin respuesta"))


def package_show(dataset: str) -> dict:
    r = get(BA_DATA_API, params={"id": dataset})
    payload = r.json()
    if not payload.get("success") or not payload.get("result"):
        raise RuntimeError(f"CKAN no devolvió el dataset {dataset}")
    return payload["result"]


def normal_format(v: str) -> str:
    s = (v or "").lower().strip()
    if "spreadsheetml" in s or s in {"xls", "xlsx", "excel"}:
        return "xlsx"
    if "csv" in s:
        return "csv"
    return s


def choose_resource(pkg: dict, pattern: str, fmt: str) -> dict:
    rx = re.compile(pattern, re.I)
    matches = []
    for res in pkg.get("resources") or []:
        name = res.get("name") or ""
        rf = normal_format(res.get("format") or res.get("mimetype") or "")
        if rx.search(name) and (not fmt or rf == fmt):
            matches.append(res)
    if not matches:
        # Algunos recursos históricos traen format vacío pero el nombre sigue siendo estable.
        for res in pkg.get("resources") or []:
            if rx.search(res.get("name") or ""):
                matches.append(res)
    if not matches:
        disponibles = [r.get("name") for r in pkg.get("resources") or []]
        raise RuntimeError(f"recurso no encontrado; disponibles: {disponibles}")
    # CKAN deja recursos históricos; se prioriza el que tenga last_modified más reciente.
    matches.sort(key=lambda r: r.get("last_modified") or r.get("created") or "", reverse=True)
    return matches[0]


def validate(path: Path, fmt: str) -> None:
    if path.stat().st_size < 500:
        raise RuntimeError("archivo demasiado chico")
    if fmt == "xlsx":
        # Se abre como stream: openpyxl no debe depender de la extensión del
        # archivo temporal (.nuevo.xlsx). Así validamos antes de reemplazar la
        # copia buena sin provocar InvalidFileException por el nombre temporal.
        with path.open("rb") as fh:
            wb = load_workbook(fh, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            wb.close()
        if len([h for h in headers if h]) < 4:
            raise RuntimeError("XLSX sin encabezados reconocibles")
    elif fmt == "csv":
        raw = path.read_bytes()[:4096]
        if b"\x00" in raw:
            raise RuntimeError("CSV parece binario")
        text = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                pass
        if not text or len(text.splitlines()) < 2:
            raise RuntimeError("CSV sin filas")


def load_state() -> dict:
    try:
        return json.loads(ESTADO.read_text(encoding="utf-8"))
    except Exception:
        return {"datasets": {}}


def save_state(state: dict) -> None:
    ESTADO.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    state = load_state()
    state.setdefault("datasets", {})
    ok, kept, failed = [], [], []
    for key, cfg in DATASETS_TERRITORIO.items():
        dst = DIR / cfg["filename"]
        existed = dst.exists()
        print(f"… {key:18} {cfg['dataset']}")
        try:
            pkg = package_show(cfg["dataset"])
            res = choose_resource(pkg, cfg["resource_pattern"], cfg["format"])
            url = res.get("url")
            if not url:
                raise RuntimeError("recurso sin URL")
            r = get(url)
            # Mantener la extensión real al final del temporal facilita además
            # la inspección manual de los artefactos durante una corrida fallida.
            tmp = dst.with_name(f"{dst.stem}.nuevo{dst.suffix}")
            tmp.write_bytes(r.content)
            try:
                validate(tmp, cfg["format"])
            except Exception:
                tmp.unlink(missing_ok=True)
                raise
            os.replace(tmp, dst)
            state["datasets"][key] = {
                "dataset": cfg["dataset"],
                "resource_id": res.get("id"),
                "resource_name": res.get("name"),
                "resource_url": url,
                "source_last_modified": res.get("last_modified") or pkg.get("metadata_modified"),
                "downloaded_at": int(time.time()),
                "bytes": dst.stat().st_size,
            }
            ok.append(key)
            print(f"  ✔ {dst.name} · {dst.stat().st_size//1024} KB")
        except Exception as e:
            if existed:
                kept.append(key)
                print(f"  ~ se conserva la copia anterior ({type(e).__name__}: {e})")
            else:
                failed.append(key)
                print(f"  ✘ sin copia local ({type(e).__name__}: {e})")
    state["updated_at"] = int(time.time())
    save_state(state)
    print(f"\nTerritorio: descargados {len(ok)} · conservados {len(kept)} · sin copia {len(failed)}")
    # No corta: verificar_territorio.py decide si el resultado puede publicarse.
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
