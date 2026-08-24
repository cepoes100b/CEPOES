#!/usr/bin/env python3
from __future__ import annotations

import io
import re
from pathlib import Path

import requests
from openpyxl import load_workbook

OUT = Path("diagnostico_fuentes_estructura_actual.txt")
SOURCES = {
    "OEDE_empresas": "https://www.argentina.gob.ar/sites/default/files/provinciales_serie_empresas1_2.xlsx",
    "IDECBA_ejes": "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2026/06/AC_EJ_2026_08.xlsx",
}
TIMEOUT = 180
TERMS = (
    "ciudad autonoma", "ciudad de buenos aires", "capital federal", "caba",
    "comuna", "indumentaria", "alimentos", "empresas", "rama", "2025", "2026",
)


def clean(v):
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v).strip())
    return s


def norm(v):
    s = clean(v).lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")):
        s = s.replace(a, b)
    return s


def fmt_row(values, max_cols=20):
    vals = [clean(v) for v in values[:max_cols]]
    while vals and not vals[-1]:
        vals.pop()
    return " | ".join(vals)


def inspect_book(name: str, url: str) -> list[str]:
    lines = [f"=== {name} ===", f"URL: {url}"]
    r = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": "CEPOES-data/1.0"})
    r.raise_for_status()
    lines.append(f"Descarga: {len(r.content)/1024/1024:.2f} MB · {r.headers.get('content-type')}")
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    lines.append("Hojas: " + ", ".join(wb.sheetnames))
    for ws in wb.worksheets:
        lines.append("")
        lines.append(f"-- Hoja: {ws.title} · {ws.max_row} filas × {ws.max_column} columnas --")
        shown = 0
        hits = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = list(row)
            text = " ".join(norm(v) for v in vals if v is not None)
            if text and shown < 35:
                lines.append(f"R{i}: {fmt_row(vals)}")
                shown += 1
            if text and any(t in text for t in TERMS) and len(hits) < 80:
                hits.append((i, fmt_row(vals, 30)))
            if i >= 5000 and len(hits) >= 20:
                break
        if hits:
            lines.append("Coincidencias relevantes:")
            for i, txt in hits:
                lines.append(f"HIT R{i}: {txt}")
    return lines


def main():
    out = []
    for name, url in SOURCES.items():
        try:
            out.extend(inspect_book(name, url))
        except Exception as e:
            out.extend([f"=== {name} ===", f"ERROR: {type(e).__name__}: {e}"])
        out.append("")
    OUT.write_text("\n".join(out), encoding="utf-8")
    print(OUT.read_text(encoding="utf-8")[:10000])


if __name__ == "__main__":
    main()
