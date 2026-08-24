#!/usr/bin/env python3
from __future__ import annotations

import io
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

OUT = Path("deploy/site-overlay/assets/data/estructura-productiva/actual.json")
TIMEOUT = 180
UA = {"User-Agent": "CEPOES-data/1.0 (+https://cepoes.org/)"}

OEDE_URL = "https://www.argentina.gob.ar/sites/default/files/provinciales_serie_empresas1_2.xlsx"
IDECBA_INDEX = "https://www.estadisticaciudad.gob.ar/eyc/categoria-banco-datos/ejes-comerciales/"
IDECBA_RUBRO_FALLBACK = "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2026/06/AC_EJ_2026_08.xlsx"
IDECBA_INDICADORES_FALLBACK = "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2026/06/AC_EJ_2026_04.xlsx"


def clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def norm(v) -> str:
    s = unicodedata.normalize("NFKD", clean(v)).encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean(v).replace("%", "").replace("\u00a0", " ")
    if not s or norm(s) in {"s.d.", "s/d", "sd", "-", "..."}:
        return None
    s = s.replace(".", "").replace(",", ".") if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*(?:,\d+)?", s) else s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def as_int(v):
    n = num(v)
    return int(round(n)) if n is not None else None


def get(url: str) -> bytes:
    r = requests.get(url, timeout=TIMEOUT, headers=UA)
    r.raise_for_status()
    return r.content


def get_text(url: str) -> str:
    r = requests.get(url, timeout=TIMEOUT, headers=UA)
    r.raise_for_status()
    return r.text


def hrefs(html: str, base: str) -> list[str]:
    out = []
    for raw in re.findall(r'''href=["']([^"']+)["']''', html, flags=re.I):
        u = urljoin(base, raw.replace("&amp;", "&"))
        if u.startswith("https://www.estadisticaciudad.gob.ar/"):
            out.append(u)
    return list(dict.fromkeys(out))


def discover_idecba_xlsx() -> tuple[list[str], list[str]]:
    """Devuelve candidatos rubro/comuna e indicadores/comuna.

    El índice se inspecciona cada ejecución para que una nueva publicación
    cuatrimestral pueda reemplazar los fallbacks sin cambiar código.
    """
    rubro = [IDECBA_RUBRO_FALLBACK]
    indicadores = [IDECBA_INDICADORES_FALLBACK]
    try:
        index_html = get_text(IDECBA_INDEX)
        article_links = [
            u for u in hrefs(index_html, IDECBA_INDEX)
            if "/banco-datos/" in u and (
                "locales-ocupados-por-comuna-segun-rubro" in u
                or "locales-relevados-ocupados-densidad-comercial" in u
            )
        ][:30]
        for page in article_links:
            try:
                html = get_text(page)
            except Exception:
                continue
            files = [u for u in hrefs(html, page) if re.search(r"\.xlsx(?:\?|$)", u, re.I)]
            if "locales-ocupados-por-comuna-segun-rubro" in page:
                rubro.extend(files)
            elif "locales-relevados-ocupados-densidad-comercial" in page and "por-comuna" in page:
                indicadores.extend(files)
    except Exception as e:
        print(f"IDECBA: descubrimiento web no disponible; uso fallbacks validados: {e}")
    return list(dict.fromkeys(rubro)), list(dict.fromkeys(indicadores))


def period_from_text(text: str) -> tuple[int, int] | None:
    s = norm(text)
    year_m = re.findall(r"(?:19|20)\d{2}", s)
    if not year_m:
        return None
    year = max(map(int, year_m))
    q = 0
    if re.search(r"\b1(?:er|ro)?\.?\s*cuatr", s): q = 1
    if re.search(r"\b2(?:do)?\.?\s*cuatr", s): q = max(q, 2)
    if re.search(r"\b3(?:er|ro)?\.?\s*cuatr", s): q = max(q, 3)
    return year, q


def workbook_latest_period(wb) -> tuple[int, int]:
    periods = []
    for name in wb.sheetnames:
        p = period_from_text(name)
        if p: periods.append(p)
    if not periods:
        for ws in wb.worksheets[:2]:
            for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True):
                p = period_from_text(" ".join(clean(v) for v in row if v is not None))
                if p: periods.append(p)
    return max(periods) if periods else (0, 0)


def select_latest_book(urls: list[str], label: str):
    best = None
    errors = []
    for url in urls:
        try:
            raw = get(url)
            wb = load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
            p = workbook_latest_period(wb)
            print(f"{label}: {url} -> período {p}")
            if best is None or p > best[0]:
                best = (p, url, wb)
        except Exception as e:
            errors.append(f"{url}: {type(e).__name__}: {e}")
    if best is None:
        raise RuntimeError(f"No pude abrir ningún XLSX de {label}: {' | '.join(errors)}")
    if best[0][0] < 2026:
        raise RuntimeError(f"{label}: el último período descubierto es {best[0]}, demasiado antiguo")
    return best


def latest_period_sheet(wb):
    candidates = []
    for ws in wb.worksheets:
        p = period_from_text(ws.title)
        if p:
            candidates.append((p, ws))
    if not candidates:
        raise RuntimeError("No encontré hojas por cuatrimestre")
    return max(candidates, key=lambda x: x[0])


def parse_idecba_rubro(wb):
    period, ws = latest_period_sheet(wb)
    # Busca la fila con las 15 comunas consecutivas.
    header_row = None
    comuna_cols = {}
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [as_int(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 40) + 1)]
        found = {v: i + 1 for i, v in enumerate(vals) if v is not None and 1 <= v <= 15}
        if len(found) >= 15:
            header_row = r
            comuna_cols = {i: found[i] for i in range(1, 16)}
            break
    if not header_row:
        raise RuntimeError(f"IDECBA rubros: no pude detectar columnas de las 15 comunas en {ws.title}")

    label_col = 1
    total_col = 2
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = clean(ws.cell(r, label_col).value)
        if not label:
            continue
        nlabel = norm(label)
        if nlabel.startswith("fuente") or nlabel.startswith("nota"):
            break
        total = as_int(ws.cell(r, total_col).value)
        vals = {str(c): as_int(ws.cell(r, col).value) or 0 for c, col in comuna_cols.items()}
        if total is None:
            continue
        rows.append({"rubro": label, "total": total, "comunas": vals})

    total_row = next((x for x in rows if norm(x["rubro"]) == "total"), None)
    if not total_row or total_row["total"] < 5000:
        raise RuntimeError("IDECBA rubros: total de locales ocupados inválido")
    rubros = [x for x in rows if norm(x["rubro"]) != "total"]
    if len(rubros) < 8:
        raise RuntimeError(f"IDECBA rubros: sólo {len(rubros)} rubros")
    return {
        "periodo": {"anio": period[0], "cuatrimestre": period[1], "etiqueta": ws.title},
        "locales_ocupados": total_row["total"],
        "ocupados_por_comuna": total_row["comunas"],
        "rubros": rubros,
    }


def expand_headers(ws, last_header_row: int) -> dict[int, str]:
    # Propaga valores de celdas combinadas horizontal y verticalmente.
    matrix = [[clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)] for r in range(1, last_header_row + 1)]
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= last_header_row:
            value = clean(ws.cell(rng.min_row, rng.min_col).value)
            if value:
                for r in range(rng.min_row, min(rng.max_row, last_header_row) + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        matrix[r - 1][c - 1] = value
    return {c: " | ".join(x for x in (matrix[r - 1][c - 1] for r in range(1, last_header_row + 1)) if x) for c in range(1, ws.max_column + 1)}


def detect_commune_rows(ws):
    for c in range(1, min(ws.max_column, 12) + 1):
        hits = []
        for r in range(1, ws.max_row + 1):
            v = as_int(ws.cell(r, c).value)
            if v is not None and 1 <= v <= 15:
                hits.append((r, v))
        for i in range(len(hits)):
            seq = hits[i:i + 15]
            if len(seq) == 15 and [v for _, v in seq] == list(range(1, 16)):
                return c, [r for r, _ in seq]
    raise RuntimeError(f"IDECBA indicadores: no pude detectar filas de comunas en {ws.title}")


def choose_col(headers: dict[int, str], needles: list[str], period: tuple[int, int], exclude: tuple[str, ...] = ()):
    candidates = []
    for c, text in headers.items():
        n = norm(text)
        if all(needle in n for needle in needles) and not any(x in n for x in exclude):
            score = 0
            if str(period[0]) in n: score += 5
            if period[1] == 1 and "1er" in n: score += 2
            if period[1] == 2 and "2do" in n: score += 2
            if period[1] == 3 and "3er" in n: score += 2
            score += c / 1000
            candidates.append((score, c))
    return max(candidates)[1] if candidates else None


def parse_idecba_indicadores(wb):
    period, ws = latest_period_sheet(wb)
    comuna_col, rows = detect_commune_rows(ws)
    first_data = min(rows)
    headers = expand_headers(ws, first_data - 1)

    relevados_col = choose_col(headers, ["relevados"], period)
    ocupados_col = choose_col(headers, ["ocupados"], period, exclude=("desocupados",))
    tasa_col = choose_col(headers, ["tasa", "ocupacion"], period)
    densidad_col = choose_col(headers, ["densidad"], period)
    var_prev_col = choose_col(headers, ["variacion", "relevamiento", "previo"], period)
    var_ia_col = choose_col(headers, ["variacion", "interanual"], period)

    # Fallbacks numéricos: se usan sólo si la etiqueta del XLSX cambia.
    numeric_cols = []
    for c in range(1, ws.max_column + 1):
        vals = [num(ws.cell(r, c).value) for r in rows]
        vals = [v for v in vals if v is not None]
        if len(vals) >= 13:
            numeric_cols.append((c, vals))
    if relevados_col is None:
        cand = [(c, vals) for c, vals in numeric_cols if max(vals) > 500 and sum(vals) > 10000 and c != comuna_col]
        if cand: relevados_col = max(cand, key=lambda x: sum(x[1]))[0]
    if ocupados_col is None:
        cand = [(c, vals) for c, vals in numeric_cols if max(vals) > 300 and sum(vals) > 8000 and c not in {comuna_col, relevados_col}]
        if cand: ocupados_col = max(cand, key=lambda x: sum(x[1]))[0]
    if tasa_col is None:
        cand = [(c, vals) for c, vals in numeric_cols if min(vals) >= 60 and max(vals) <= 100.5]
        if cand: tasa_col = cand[-1][0]

    if relevados_col is None or ocupados_col is None:
        raise RuntimeError(f"IDECBA indicadores: columnas insuficientes. Headers={headers}")

    comunas = {}
    for r in rows:
        c = as_int(ws.cell(r, comuna_col).value)
        relevados = as_int(ws.cell(r, relevados_col).value)
        ocupados = as_int(ws.cell(r, ocupados_col).value)
        if c is None or relevados is None or ocupados is None:
            continue
        tasa = num(ws.cell(r, tasa_col).value) if tasa_col else None
        if tasa is not None and tasa <= 1.5:
            tasa *= 100
        if tasa is None:
            tasa = 100 * ocupados / relevados if relevados else None
        densidad = num(ws.cell(r, densidad_col).value) if densidad_col else None
        comunas[str(c)] = {
            "relevados": relevados,
            "ocupados": ocupados,
            "tasa_ocupacion": round(tasa, 1) if tasa is not None else None,
            "densidad_comercial": round(densidad, 2) if densidad is not None else None,
            "variacion_previa": round(num(ws.cell(r, var_prev_col).value), 2) if var_prev_col and num(ws.cell(r, var_prev_col).value) is not None else None,
            "variacion_interanual": round(num(ws.cell(r, var_ia_col).value), 2) if var_ia_col and num(ws.cell(r, var_ia_col).value) is not None else None,
        }
    if len(comunas) != 15:
        raise RuntimeError(f"IDECBA indicadores: {len(comunas)} comunas, se esperaban 15")
    total_relevados = sum(x["relevados"] for x in comunas.values())
    total_ocupados = sum(x["ocupados"] for x in comunas.values())
    if total_relevados < 10000 or total_ocupados < 9000:
        raise RuntimeError(f"IDECBA indicadores: totales improbables {total_relevados}/{total_ocupados}")
    return {
        "periodo": {"anio": period[0], "cuatrimestre": period[1], "etiqueta": ws.title},
        "locales_relevados": total_relevados,
        "locales_ocupados": total_ocupados,
        "tasa_ocupacion": round(100 * total_ocupados / total_relevados, 1),
        "comunas": comunas,
    }


def parse_oede(raw: bytes):
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if "Capital Federal" not in wb.sheetnames:
        raise RuntimeError("OEDE: falta hoja Capital Federal")
    ws = wb["Capital Federal"]

    year_cols = {}
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, ws.max_column + 1):
            y = as_int(ws.cell(r, c).value)
            if y is not None and 1990 <= y <= 2035:
                year_cols[y] = c
                header_row = r
        if len(year_cols) >= 10:
            break
    if not year_cols:
        raise RuntimeError("OEDE: no detecté años")
    latest = max(year_cols)
    if latest < 2024:
        raise RuntimeError(f"OEDE: último año {latest}, se esperaba al menos 2024")

    sectors = []
    section_rows = []
    for r in range((header_row or 4) + 1, ws.max_row + 1):
        code = clean(ws.cell(r, 1).value).upper()
        label = clean(ws.cell(r, 2).value)
        if re.fullmatch(r"[A-Z]", code) and label:
            section_rows.append((r, code, label))
            value = as_int(ws.cell(r, year_cols[latest]).value)
            if value is not None:
                sectors.append({"codigo": code, "sector": label.title(), "empresas": value})
    if len(sectors) < 8:
        raise RuntimeError(f"OEDE: sólo {len(sectors)} secciones económicas")

    def total_for_year(y):
        col = year_cols[y]
        vals = [as_int(ws.cell(r, col).value) for r, _, _ in section_rows]
        return sum(v for v in vals if v is not None)

    total = total_for_year(latest)
    if not 50000 <= total <= 250000:
        raise RuntimeError(f"OEDE: total anual improbable {total}")
    years = sorted(y for y in year_cols if y >= 2015)
    serie = [{"anio": y, "empresas": total_for_year(y)} for y in years]
    sectors.sort(key=lambda x: x["empresas"], reverse=True)
    return {
        "periodo": latest,
        "empresas": total,
        "sectores": sectors,
        "serie": serie,
        "nota": "Empresas privadas con empleo asalariado registrado. Una firma puede contabilizarse en más de una provincia si declara personal en distintas jurisdicciones.",
    }


def main():
    rubro_urls, indicadores_urls = discover_idecba_xlsx()
    rubro_period, rubro_url, rubro_wb = select_latest_book(rubro_urls, "IDECBA rubros/comuna")
    ind_period, ind_url, ind_wb = select_latest_book(indicadores_urls, "IDECBA indicadores/comuna")
    if rubro_period != ind_period:
        raise RuntimeError(f"IDECBA: períodos no coinciden: rubros={rubro_period}, indicadores={ind_period}")

    rubro = parse_idecba_rubro(rubro_wb)
    indicadores = parse_idecba_indicadores(ind_wb)
    # Control cruzado entre los dos tabulados oficiales.
    if abs(rubro["locales_ocupados"] - indicadores["locales_ocupados"]) > 2:
        raise RuntimeError(f"IDECBA: ocupados no coinciden: {rubro['locales_ocupados']} vs {indicadores['locales_ocupados']}")
    for c in range(1, 16):
        a = rubro["ocupados_por_comuna"][str(c)]
        b = indicadores["comunas"][str(c)]["ocupados"]
        if abs(a - b) > 2:
            raise RuntimeError(f"IDECBA: comuna {c}, ocupados no coinciden: {a} vs {b}")

    oede = parse_oede(get(OEDE_URL))

    obj = {
        "schema": 1,
        "generado": datetime.now(timezone.utc).isoformat(),
        "panorama": {
            "empresas_registradas": oede,
            "ejes_comerciales": {
                **indicadores,
                "rubros": rubro["rubros"],
                "universo": "48 ejes comerciales de alta densidad; no representa la totalidad de los locales de CABA.",
            },
        },
        "fuentes": {
            "oede": {
                "nombre": "Observatorio de Empleo y Dinámica Empresarial (OEDE) · SIPA",
                "url": OEDE_URL,
                "unidad": "empresa privada con empleo asalariado registrado",
                "periodo": oede["periodo"],
            },
            "idecba_rubros": {
                "nombre": "IDECBA · Locales ocupados por comuna según rubro · 48 ejes comerciales",
                "url": rubro_url,
                "unidad": "local comercial ocupado",
                "periodo": rubro["periodo"],
            },
            "idecba_indicadores": {
                "nombre": "IDECBA · Locales relevados, ocupados, densidad y tasa de ocupación por comuna · 48 ejes comerciales",
                "url": ind_url,
                "unidad": "local comercial relevado/ocupado",
                "periodo": indicadores["periodo"],
            },
        },
        "criterio": {
            "actualidad": "La vista principal usa el último dato oficial disponible de cada universo. No se interpola RUS 2017 para estimar un stock 2026.",
            "unidades": "Empresa registrada, local comercial ocupado, habilitación aprobada y establecimiento RUS son unidades distintas y se muestran por separado.",
            "historico": "El RUS 2017 se conserva exclusivamente como capa histórica de alta resolución territorial hasta nivel manzana.",
        },
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(obj, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"{OUT} · {OUT.stat().st_size/1024:.1f} KB")
    print(f"OEDE {oede['periodo']}: {oede['empresas']:,} empresas")
    print(f"IDECBA {rubro_period}: {indicadores['locales_ocupados']:,}/{indicadores['locales_relevados']:,} locales ocupados/relevados · {indicadores['tasa_ocupacion']}%")


if __name__ == "__main__":
    main()
