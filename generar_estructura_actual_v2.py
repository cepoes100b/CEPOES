#!/usr/bin/env python3
from __future__ import annotations

import io
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

OUT = Path("deploy/site-overlay/assets/data/estructura-productiva/actual.json")
TIMEOUT = 180
UA = {"User-Agent": "CEPOES-data/1.0 (+https://cepoes.org/)"}
OEDE_URL = "https://www.argentina.gob.ar/sites/default/files/provinciales_serie_empresas1_2.xlsx"
IDECBA_INDEX = "https://www.estadisticaciudad.gob.ar/eyc/categoria-banco-datos/ejes-comerciales/"
RUBRO_FALLBACK = "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2026/06/AC_EJ_2026_08.xlsx"
IND_FALLBACK = "https://www.estadisticaciudad.gob.ar/eyc/wp-content/uploads/2026/06/AC_EJ_2026_04.xlsx"


def clean(v):
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def norm(v):
    return unicodedata.normalize("NFKD", clean(v)).encode("ascii", "ignore").decode().lower().strip()


def number(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean(v).replace("%", "")
    if not s or norm(s) in {"s.d.", "s/d", "sd", "-", "..."}:
        return None
    if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*(?:,\d+)?", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def as_int(v):
    n = number(v)
    return int(round(n)) if n is not None else None


def get(url):
    r = requests.get(url, timeout=TIMEOUT, headers=UA)
    r.raise_for_status()
    return r.content


def get_text(url):
    r = requests.get(url, timeout=TIMEOUT, headers=UA)
    r.raise_for_status()
    return r.text


def hrefs(html, base):
    out = []
    for raw in re.findall(r'''href=["']([^"']+)["']''', html, re.I):
        u = urljoin(base, raw.replace("&amp;", "&"))
        if u.startswith("https://www.estadisticaciudad.gob.ar/"):
            out.append(u)
    return list(dict.fromkeys(out))


def period(text):
    s = norm(text)
    years = [int(x) for x in re.findall(r"(?:19|20)\d{2}", s)]
    if not years:
        return None
    q = 0
    if re.search(r"\b1(?:er|ro)?\.?\s*cuatr", s): q = 1
    if re.search(r"\b2(?:do)?\.?\s*cuatr", s): q = 2
    if re.search(r"\b3(?:er|ro)?\.?\s*cuatr", s): q = 3
    return max(years), q


def discover():
    rubros = [RUBRO_FALLBACK]
    indicadores = [IND_FALLBACK]
    try:
        html = get_text(IDECBA_INDEX)
        pages = [u for u in hrefs(html, IDECBA_INDEX) if "/banco-datos/" in u]
        for page in pages[:50]:
            np = norm(page)
            if "locales-ocupados-por-comuna-segun-rubro" not in np and "locales-relevados-ocupados-densidad-comercial" not in np:
                continue
            try:
                files = [u for u in hrefs(get_text(page), page) if re.search(r"\.xlsx(?:\?|$)", u, re.I)]
            except Exception:
                continue
            if "locales-ocupados-por-comuna-segun-rubro" in np:
                rubros.extend(files)
            elif "por-comuna" in np:
                indicadores.extend(files)
    except Exception as e:
        print("IDECBA discovery fallback:", e)
    return list(dict.fromkeys(rubros)), list(dict.fromkeys(indicadores))


def latest_book(urls, label):
    best = None
    for url in urls:
        try:
            wb = load_workbook(io.BytesIO(get(url)), read_only=False, data_only=True)
            ps = [period(s) for s in wb.sheetnames if period(s)]
            p = max(ps) if ps else (0, 0)
            print(f"{label}: {url} -> {p}")
            if best is None or p > best[0]: best = (p, url, wb)
        except Exception as e:
            print(f"{label}: descarto {url}: {e}")
    if best is None or best[0][0] < 2026:
        raise RuntimeError(f"{label}: no hay libro 2026 válido")
    return best


def latest_sheet(wb):
    xs = [(period(ws.title), ws) for ws in wb.worksheets if period(ws.title)]
    if not xs:
        raise RuntimeError("IDECBA: no hay hoja por cuatrimestre")
    return max(xs, key=lambda x: x[0])


def parse_rubros(wb):
    p, ws = latest_sheet(wb)
    header = None
    cols = None
    for r in range(1, min(15, ws.max_row) + 1):
        row = [as_int(ws.cell(r, c).value) for c in range(1, min(45, ws.max_column) + 1)]
        positions = {}
        for c, v in enumerate(row, 1):
            if v is not None and 1 <= v <= 15 and v not in positions:
                positions[v] = c
        if len(positions) == 15:
            header, cols = r, positions
            break
    if not header:
        raise RuntimeError("IDECBA rubros: no detecté las 15 comunas")
    out = []
    for r in range(header + 1, ws.max_row + 1):
        label = clean(ws.cell(r, 1).value)
        if not label: continue
        if norm(label).startswith(("fuente", "nota")): break
        total = as_int(ws.cell(r, 2).value)
        if total is None: continue
        out.append({"rubro": label, "total": total, "comunas": {str(c): as_int(ws.cell(r, cols[c]).value) or 0 for c in range(1, 16)}})
    totalrow = next((x for x in out if norm(x["rubro"]) == "total"), None)
    if not totalrow or totalrow["total"] < 9000:
        raise RuntimeError("IDECBA rubros: total inválido")
    rubros = [x for x in out if norm(x["rubro"]) != "total"]
    if len(rubros) < 8 or sum(x["total"] for x in rubros) != totalrow["total"]:
        raise RuntimeError("IDECBA rubros: composición inconsistente")
    return p, totalrow["total"], totalrow["comunas"], rubros


def commune_marker(v):
    s = norm(v)
    m = re.fullmatch(r"comuna\s*(\d{1,2})", s)
    if m and 1 <= int(m.group(1)) <= 15:
        return int(m.group(1)), True
    n = as_int(v)
    if n is not None and 1 <= n <= 15:
        return n, False
    return None, False


def parse_indicadores(wb, occupied_by_comuna):
    p, ws = latest_sheet(wb)
    candidates = {i: [] for i in range(1, 16)}
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        markers = [(commune_marker(v), c) for c, v in enumerate(vals, 1)]
        markers = [(m, c) for m, c in markers if m[0] is not None]
        nums = [(c, number(v)) for c, v in enumerate(vals, 1) if number(v) is not None]
        for (comuna, explicit), marker_col in markers:
            target = occupied_by_comuna[str(comuna)]
            if target <= 0: continue
            occ_hits = [(c, n) for c, n in nums if abs(n - target) <= 1]
            if not occ_hits: continue
            relev = [n for c, n in nums if n > target and n <= target * 1.25 and abs(n - round(n)) < 1e-8]
            if not relev: continue
            relevados = int(round(min(relev)))
            score = (10 if explicit else 0) + (5 if marker_col <= 3 else 0) + len(occ_hits)
            candidates[comuna].append((score, r, relevados))

    comunas = {}
    missing = []
    for c in range(1, 16):
        if not candidates[c]:
            missing.append(c)
            continue
        score, r, relevados = max(candidates[c], key=lambda x: (x[0], -x[1]))
        ocupados = occupied_by_comuna[str(c)]
        tasa = round(100 * ocupados / relevados, 1)
        if not 75 <= tasa <= 100:
            raise RuntimeError(f"IDECBA comuna {c}: tasa improbable {tasa}; fila {r}")
        comunas[str(c)] = {"relevados": relevados, "ocupados": ocupados, "tasa_ocupacion": tasa}
    if missing:
        preview = []
        for r in range(1, min(ws.max_row, 35) + 1):
            row = [clean(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 20) + 1)]
            if any(row): preview.append(f"R{r}: {' | '.join(row)}")
        raise RuntimeError("IDECBA indicadores: faltan comunas " + str(missing) + "\n" + "\n".join(preview))
    relevados = sum(x["relevados"] for x in comunas.values())
    ocupados = sum(x["ocupados"] for x in comunas.values())
    if not 12000 <= relevados <= 14000 or not 11000 <= ocupados <= 12500:
        raise RuntimeError(f"IDECBA indicadores: totales improbables {relevados}/{ocupados}")
    return p, relevados, ocupados, round(100 * ocupados / relevados, 1), comunas


def parse_oede(raw):
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["Capital Federal"]
    years = {}
    header = None
    for r in range(1, 10):
        found = {}
        for c in range(1, ws.max_column + 1):
            y = as_int(ws.cell(r, c).value)
            if y and 1990 <= y <= 2035: found[y] = c
        if len(found) >= 20:
            years, header = found, r
            break
    if not years:
        raise RuntimeError("OEDE: no detecté serie anual")
    latest = max(years)
    if latest < 2024:
        raise RuntimeError(f"OEDE: último año {latest}")
    rows = []
    for r in range(header + 1, ws.max_row + 1):
        code = clean(ws.cell(r, 1).value).upper()
        label = clean(ws.cell(r, 2).value)
        if re.fullmatch(r"[A-Z]", code) and label:
            rows.append((r, code, label))
    if not 8 <= len(rows) <= 25:
        raise RuntimeError(f"OEDE: {len(rows)} secciones; posible tabla duplicada")

    def total(y):
        return sum(as_int(ws.cell(r, years[y]).value) or 0 for r, _, _ in rows)

    total_latest = total(latest)
    if not 80000 <= total_latest <= 180000:
        raise RuntimeError(f"OEDE: total {latest} improbable: {total_latest}")
    sectors = [{"codigo": code, "sector": label.title(), "empresas": as_int(ws.cell(r, years[latest]).value) or 0} for r, code, label in rows]
    sectors.sort(key=lambda x: x["empresas"], reverse=True)
    return {
        "periodo": latest,
        "empresas": total_latest,
        "sectores": sectors,
        "serie": [{"anio": y, "empresas": total(y)} for y in sorted(y for y in years if y >= 2015)],
        "nota": "Empresas privadas con empleo asalariado registrado; una firma puede contabilizarse en más de una jurisdicción si declara personal en distintas provincias.",
    }


def main():
    rubro_urls, ind_urls = discover()
    rp, rubro_url, rubro_wb = latest_book(rubro_urls, "IDECBA rubros")
    ip, ind_url, ind_wb = latest_book(ind_urls, "IDECBA indicadores")
    if rp != ip:
        raise RuntimeError(f"IDECBA: períodos distintos {rp} / {ip}")
    _, ocupados_rubro, ocupados_comuna, rubros = parse_rubros(rubro_wb)
    _, relevados, ocupados, tasa, comunas = parse_indicadores(ind_wb, ocupados_comuna)
    if ocupados != ocupados_rubro:
        raise RuntimeError(f"IDECBA: ocupados no coinciden {ocupados} / {ocupados_rubro}")
    oede = parse_oede(get(OEDE_URL))

    d = {
        "schema": 1,
        "generado": datetime.now(timezone.utc).isoformat(),
        "panorama": {
            "empresas_registradas": oede,
            "ejes_comerciales": {
                "periodo": {"anio": rp[0], "cuatrimestre": rp[1]},
                "locales_relevados": relevados,
                "locales_ocupados": ocupados,
                "tasa_ocupacion": tasa,
                "comunas": comunas,
                "rubros": rubros,
                "universo": "48 ejes comerciales de alta densidad; no representa la totalidad de los locales de CABA."
            }
        },
        "fuentes": {
            "oede": {"nombre": "OEDE · SIPA", "url": OEDE_URL, "unidad": "empresa privada con empleo asalariado registrado", "periodo": oede["periodo"]},
            "idecba_rubros": {"nombre": "IDECBA · Locales ocupados por comuna según rubro · 48 ejes comerciales", "url": rubro_url, "unidad": "local comercial ocupado", "periodo": {"anio": rp[0], "cuatrimestre": rp[1]}},
            "idecba_indicadores": {"nombre": "IDECBA · Locales relevados y ocupados por comuna · 48 ejes comerciales", "url": ind_url, "unidad": "local comercial relevado/ocupado", "periodo": {"anio": ip[0], "cuatrimestre": ip[1]}}
        },
        "criterio": {
            "actualidad": "La vista principal usa el último dato oficial disponible de cada universo. No se interpola RUS 2017 para estimar un stock 2026.",
            "unidades": "Empresa registrada, local comercial ocupado, habilitación aprobada y establecimiento RUS son unidades distintas y se muestran por separado.",
            "historico": "El RUS 2017 se conserva exclusivamente como capa histórica de alta resolución territorial hasta nivel manzana."
        }
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(d, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"{OUT} · OEDE {oede['periodo']}: {oede['empresas']:,} empresas · IDECBA {rp}: {ocupados:,}/{relevados:,} · {tasa}%")


if __name__ == "__main__":
    main()
