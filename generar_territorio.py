"""Genera territorio.json con equipamientos y espacio público por comuna/barrio.

Fuentes: BA Data. La población usada para m² de espacio verde por habitante sale
del bloque Censo 2022 ya validado en datos.json.
"""
from __future__ import annotations

import csv
import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from fuentes_territorio import BA_DATA_BASE, DATASETS_TERRITORIO

BASE = Path(__file__).resolve().parent
DIR = BASE / "badata"
DATOS = BASE / "datos.json"
OUT = BASE / "territorio.json"
STATE = BASE / "estado_territorio.json"

BARRIO_ALIASES = {
    "paternal": "La Paternal",
    "la paternal": "La Paternal",
    "villa gral mitre": "Villa General Mitre",
    "villa gral. mitre": "Villa General Mitre",
    "villa general mitre": "Villa General Mitre",
    "montserrat": "Monserrat",
    "monserrat": "Monserrat",
    "villa pueyrredon": "Villa Pueyrredón",
}


def clean_text(v):
    if v is None:
        return ""
    s = str(v).strip()
    if "Ã" in s or "Â" in s:
        try:
            s = s.encode("latin-1").decode("utf-8")
        except Exception:
            pass
    return re.sub(r"\s+", " ", s).strip()


def norm(v):
    s = clean_text(v)
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return re.sub(r"\s+", " ", s)


def slug(v):
    return norm(v).replace(" ", "-")


def parse_comuna(v):
    s = clean_text(v)
    m = re.search(r"\b(1[0-5]|[1-9])\b", s)
    return int(m.group(1)) if m else None


def num(v):
    """Convierte números publicados con formato AR o internacional.

    BA Data mezcla celdas numéricas de XLSX (que openpyxl ya entrega como
    float) con strings. Si hay coma y punto, se interpreta punto de miles y
    coma decimal; con sólo coma, la coma es decimal; con sólo punto se respeta
    el punto decimal.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v) if math.isfinite(float(v)) else None
    s = clean_text(v).replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        out = float(s)
        return out if math.isfinite(out) else None
    except ValueError:
        return None


def canonical_barrio(v, official_by_norm):
    n = norm(v)
    if not n:
        return ""
    if n in BARRIO_ALIASES:
        return BARRIO_ALIASES[n]
    return official_by_norm.get(n, clean_text(v).title())


def read_csv(path: Path):
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(f"no pude decodificar {path.name}")
    sample = text[:10000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return list(csv.DictReader(text.splitlines(), dialect=dialect))


def read_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(x) for x in next(rows)]
    out = []
    for row in rows:
        if not any(v not in (None, "") for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    wb.close()
    return out


def pick(row, *names):
    lookup = {norm(k): v for k, v in row.items()}
    for name in names:
        n = norm(name)
        if n in lookup and lookup[n] not in (None, ""):
            return lookup[n]
    return None


def base_output(datos):
    censo = datos["censo"]
    official_by_norm = {}
    barrios = {}
    for cid, c in censo["comunas"].items():
        for name in c.get("barrios", {}):
            display = BARRIO_ALIASES.get(norm(name), name)
            official_by_norm[norm(display)] = display
            barrios[slug(display)] = {
                "nombre": display,
                "comuna": int(cid),
                "poblacion": c["barrios"][name],
                "educacion": {"establecimientos": 0, "estatales": 0, "privados": 0},
                "salud": {"hospitales": 0, "cesac": 0},
                "espacio_verde": {"espacios": 0, "m2": 0.0, "m2_hab": 0.0},
            }
    comunas = {}
    for cid, c in censo["comunas"].items():
        comunas[cid] = {
            "poblacion": c["pob"],
            "educacion": {"establecimientos": 0, "estatales": 0, "privados": 0},
            "salud": {"hospitales": 0, "cesac": 0},
            "espacio_verde": {"espacios": 0, "m2": 0.0, "m2_hab": 0.0},
        }
    return comunas, barrios, official_by_norm


def inc(communes, barrios, comuna, barrio, section, key, amount=1):
    if comuna and str(comuna) in communes:
        communes[str(comuna)][section][key] += amount
    if barrio:
        b = barrios.get(slug(barrio))
        if b:
            b[section][key] += amount


def main() -> int:
    datos = json.loads(DATOS.read_text(encoding="utf-8"))
    comunas, barrios, official_by_norm = base_output(datos)
    state = json.loads(STATE.read_text(encoding="utf-8")) if STATE.exists() else {"datasets": {}}
    errors = []

    # Educación: se cuentan sedes/establecimientos por CUE+anexo; CUI queda como fallback.
    try:
        rows = read_csv(DIR / DATASETS_TERRITORIO["educacion"]["filename"])
        seen = set()
        for r in rows:
            estado = norm(pick(r, "estado", "estado_est", "estado_est_desc", "estado_loc_desc"))
            if estado in {"2", "3", "4"} or "inactiv" in estado or "baja" in estado:
                continue
            ident = clean_text(pick(r, "cueanexo", "cue", "cui", "OBJECTID"))
            comuna = parse_comuna(pick(r, "comuna"))
            barrio = canonical_barrio(pick(r, "barrio"), official_by_norm)
            if not ident:
                ident = f"{comuna}|{barrio}|{clean_text(pick(r,'nombre_est','nombre_loc'))}"
            key = (ident, comuna, norm(barrio))
            if key in seen:
                continue
            seen.add(key)
            sector = norm(pick(r, "sector", "sector_desc"))
            inc(comunas, barrios, comuna, barrio, "educacion", "establecimientos")
            if sector in {"1", "estatal", "publico", "publica"} or "estatal" in sector:
                inc(comunas, barrios, comuna, barrio, "educacion", "estatales")
            elif sector in {"2", "privado", "privada"} or "privad" in sector:
                inc(comunas, barrios, comuna, barrio, "educacion", "privados")
    except Exception as e:
        errors.append(f"educacion: {e}")

    # CeSAC.
    try:
        rows = read_csv(DIR / DATASETS_TERRITORIO["cesac"]["filename"])
        seen = set()
        for r in rows:
            ident = clean_text(pick(r, "id", "nombre"))
            if not ident or ident in seen:
                continue
            seen.add(ident)
            comuna = parse_comuna(pick(r, "comuna"))
            barrio = canonical_barrio(pick(r, "barrio"), official_by_norm)
            inc(comunas, barrios, comuna, barrio, "salud", "cesac")
    except Exception as e:
        errors.append(f"cesac: {e}")

    # Hospitales: el XLSX vigente trae bar/com.
    try:
        rows = read_xlsx(DIR / DATASETS_TERRITORIO["hospitales"]["filename"])
        seen = set()
        for r in rows:
            name = clean_text(pick(r, "fna", "NOMBRE", "nombre", "gna"))
            if not name or norm(name) in seen:
                continue
            seen.add(norm(name))
            comuna = parse_comuna(pick(r, "com", "COMUNA", "comuna"))
            barrio = canonical_barrio(pick(r, "bar", "BARRIO", "barrio"), official_by_norm)
            inc(comunas, barrios, comuna, barrio, "salud", "hospitales")
    except Exception as e:
        errors.append(f"hospitales: {e}")

    # Espacios verdes públicos: cantidad y superficie oficial en m².
    try:
        rows = read_xlsx(DIR / DATASETS_TERRITORIO["espacios_verdes"]["filename"])
        seen = set()
        for r in rows:
            ident = clean_text(pick(r, "id", "id_ev_pub", "nombre", "nombre_ev"))
            if not ident or ident in seen:
                continue
            seen.add(ident)
            comuna = parse_comuna(pick(r, "comuna", "COMUNA"))
            barrio = canonical_barrio(pick(r, "barrio", "BARRIO"), official_by_norm)
            area = num(pick(r, "area", "SUP_TOTAL", "sup_total")) or 0.0
            inc(comunas, barrios, comuna, barrio, "espacio_verde", "espacios")
            inc(comunas, barrios, comuna, barrio, "espacio_verde", "m2", area)
    except Exception as e:
        errors.append(f"espacios_verdes: {e}")

    for cid, c in comunas.items():
        p = c["poblacion"] or 0
        c["espacio_verde"]["m2"] = round(c["espacio_verde"]["m2"], 1)
        c["espacio_verde"]["m2_hab"] = round(c["espacio_verde"]["m2"] / p, 2) if p else 0
    for b in barrios.values():
        p = b["poblacion"] or 0
        b["espacio_verde"]["m2"] = round(b["espacio_verde"]["m2"], 1)
        b["espacio_verde"]["m2_hab"] = round(b["espacio_verde"]["m2"] / p, 2) if p else 0

    sources = {}
    for key, cfg in DATASETS_TERRITORIO.items():
        st = (state.get("datasets") or {}).get(key, {})
        sources[key] = {
            "dataset": cfg["dataset"],
            "descripcion": cfg["descripcion"],
            "pagina": BA_DATA_BASE + cfg["dataset"],
            "resource_id": st.get("resource_id"),
            "source_last_modified": st.get("source_last_modified"),
        }

    out = {
        "version": 1,
        "generado": __import__("datetime").date.today().isoformat(),
        "fuente": "BA Data (Gobierno de la Ciudad de Buenos Aires)",
        "comunas": comunas,
        "barrios": barrios,
        "fuentes": sources,
        "errores": errors,
    }
    tmp = OUT.with_suffix(".json.parcial")
    tmp.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    tmp.replace(OUT)
    print(f"territorio.json · {OUT.stat().st_size//1024} KB · {len(comunas)} comunas · {len(barrios)} barrios")
    for e in errors:
        print("  ~", e)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
