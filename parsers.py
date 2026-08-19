"""Parsers de las planillas de IDECBA al formato del observatorio CEPOES.

Cada función recibe la ruta de un .xlsx y devuelve el bloque de datos.json que
le corresponde. Todas comparten el mismo contrato: si la planilla no tiene la
forma esperada, levantan ValueError con un mensaje concreto; el generador lo
captura y conserva el bloque anterior en lugar de escribir datos rotos.

Convenciones de IDECBA que hay que tolerar en todas las planillas:
  · "///", ".", "-", "s/d" y celdas vacías significan "sin dato" -> None
  · los valores pueden venir con coma decimal y con una letra de nota al pie
    pegada ("6,8a" es 6.8 con la nota "a")
  · los encabezados ocupan varias filas y las columnas de datos empiezan
    después de una fila de subtítulos
"""
import re
import unicodedata
from datetime import datetime

import openpyxl

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
MES_CORTO = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
TRIMESTRES = {"1er": 1, "2do": 2, "3er": 3, "4to": 4}
SIN_DATO = {"", "///", "//", ".", "..", "-", "--", "s/d", "sd", "n/d", "*"}


# ---------------------------------------------------------------- utilidades

def norm(v):
    """Texto normalizado: sin acentos, sin espacios dobles, en minúsculas."""
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()


def num(v):
    """Número tolerante a coma decimal, notas al pie y signos de 'sin dato'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", " ").strip()
    if norm(s) in SIN_DATO:
        return None
    s = re.sub(r"[a-zA-Z\s%$]+$", "", s).strip()      # nota al pie pegada
    s = s.replace(".", "") if s.count(",") == 1 and s.count(".") > 1 else s
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def anio_de(v):
    """Año de una celda de encabezado. IDECBA marca los años provisorios con
    asterisco ('2025*') y a veces agrega llamadas al pie, así que se limpia
    todo lo que no sea dígito antes de comparar."""
    s = re.sub(r"[^0-9]", "", str(v or ""))
    return int(s) if re.fullmatch(r"(19|20)\d{2}", s) else None


def r1(v, d=1):
    return None if v is None else round(v, d)


def abrir(path, hoja=None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if hoja is None:
        return wb, wb[wb.sheetnames[0]]
    for sn in wb.sheetnames:
        if norm(hoja) in norm(sn):
            return wb, wb[sn]
    raise ValueError(f"{path}: no encontré la hoja '{hoja}' (hay {wb.sheetnames})")


def etiqueta_mes(dt):
    return f"{MES_CORTO[dt.month - 1]}-{str(dt.year)[2:]}"


def celda_fecha(v):
    """Convierte a datetime una celda que puede ser fecha real o texto."""
    if isinstance(v, datetime):
        return v
    s = norm(v)
    m = re.match(r"(\d{4})-(\d{2})", s)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), 1)
    for nombre, k in MESES.items():
        if nombre in s:
            a = re.search(r"(19|20)\d{2}", s)
            if a:
                return datetime(int(a.group(0)), k, 1)
    return None


# ---------------------------------------------------------------- IPCBA

def ipcba(path):
    """Nivel general: índice y variación mensual. La interanual se calcula acá
    porque la planilla sólo publica la variación respecto del mes anterior."""
    _, ws = abrir(path)
    fila0 = None
    for r in range(1, 12):
        if celda_fecha(ws.cell(r, 1).value):
            fila0 = r
            break
    if fila0 is None:
        raise ValueError(f"{path}: no encontré la primera fila con fecha")

    # la columna del índice es la primera numérica; la de var. mensual viene
    # después del bloque de índices (misma cantidad de series en ambos bloques)
    meses, indices, var_m = [], [], []
    for r in range(fila0, ws.max_row + 1):
        dt = celda_fecha(ws.cell(r, 1).value)
        if dt is None:
            continue
        ix = num(ws.cell(r, 2).value)
        if ix is None:
            continue
        meses.append(dt)
        indices.append(ix)
        var_m.append(num(ws.cell(r, 6).value))

    if len(meses) < 24:
        raise ValueError(f"{path}: sólo {len(meses)} meses, esperaba 24+")

    var_ia = []
    for i, dt in enumerate(meses):
        j = next((k for k, d in enumerate(meses)
                  if d.year == dt.year - 1 and d.month == dt.month), None)
        var_ia.append(r1((indices[i] / indices[j] - 1) * 100, 1)
                      if j is not None and indices[j] else None)

    # se recorta el arranque sin interanual
    i0 = next((i for i, v in enumerate(var_ia) if v is not None), 0)
    return {
        "meses": [etiqueta_mes(d) for d in meses[i0:]],
        "var_m": [r1(v) for v in var_m[i0:]],
        "var_ia": var_ia[i0:],
    }


def ipcba_divisiones(path):
    """Índices y variaciones por división COICOP del último mes publicado."""
    _, ws = abrir(path)
    hdr = None
    for r in range(1, 12):
        etiquetas = [norm(ws.cell(r, c).value) for c in range(1, 8)]
        if any("division" in e or "nivel general" in e for e in etiquetas):
            hdr = r
            break
    if hdr is None:
        raise ValueError(f"{path}: no encontré el encabezado de divisiones")

    periodo = None
    for r in range(1, 8):
        dt = celda_fecha(ws.cell(r, 1).value) or celda_fecha(ws.cell(r, 2).value)
        if dt:
            periodo = f"{dt.year}-{dt.month:02d}"
            break

    data = []
    for r in range(hdr + 1, ws.max_row + 1):
        nombre = ws.cell(r, 1).value
        if not nombre or len(str(nombre).strip()) < 4:
            continue
        n = norm(nombre)
        if n.startswith("nota") or n.startswith("fuente") or "nivel general" in n:
            continue
        ix = num(ws.cell(r, 2).value)
        if ix is None:
            continue
        data.append({
            "nombre": str(nombre).strip(),
            "indice": r1(ix, 2),
            "var_mensual": r1(num(ws.cell(r, 3).value), 2),
            "var_ia": r1(num(ws.cell(r, 4).value), 2),
        })
    if len(data) < 8:
        raise ValueError(f"{path}: {len(data)} divisiones, esperaba 12 o 13")
    return {"periodo": periodo, "data": data}


# ---------------------------------------------------------------- canastas

def canastas(path):
    """La planilla viene transpuesta: los meses son columnas y los componentes
    filas. Se buscan las tres filas de totales por nombre."""
    _, ws = abrir(path)
    fila_meses = None
    for r in range(1, 10):
        if celda_fecha(ws.cell(r, 2).value):
            fila_meses = r
            break
    if fila_meses is None:
        raise ValueError(f"{path}: no encontré la fila de meses")

    cols, meses = [], []
    for c in range(2, ws.max_column + 1):
        dt = celda_fecha(ws.cell(fila_meses, c).value)
        if dt:
            cols.append(c)
            meses.append(dt)
    if not cols:
        raise ValueError(f"{path}: fila de meses vacía")

    claves = {"ca": "canasta alimentaria", "caysh": "canasta alimentaria y de servicios",
              "total": "canasta total"}
    out = {"meses": [f"{d.year}-{d.month:02d}" for d in meses]}
    for k, etiqueta in claves.items():
        fila = None
        for r in range(1, ws.max_row + 1):
            n = norm(ws.cell(r, 1).value)
            if n.startswith(etiqueta):
                fila = r
                break
        if fila is None:
            for r in range(1, ws.max_row + 1):
                if etiqueta.split()[1] in norm(ws.cell(r, 1).value) and "canasta" in norm(ws.cell(r, 1).value):
                    fila = r
                    break
        if fila is None:
            raise ValueError(f"{path}: no encontré la fila '{etiqueta}'")
        out[k] = [r1(num(ws.cell(fila, c).value), 2) for c in cols]
    return out


# ---------------------------------------------------------------- empleo

def empleo(path):
    """Filas de año seguidas de filas de trimestre. Cuatro tasas por trimestre."""
    _, ws = abrir(path)
    trimestres, series = [], {"actividad": [], "empleo": [],
                              "desocupacion": [], "subocupacion": []}
    anio = None
    for r in range(1, ws.max_row + 1):
        etiqueta = str(ws.cell(r, 1).value or "").strip()
        n = norm(etiqueta)
        a = anio_de(etiqueta)
        if a and not re.search(r"[a-z]", n):
            anio = a
            continue
        m = re.match(r"(1er|2do|3er|4to)\.?\s*trimestre", n)
        if not (m and anio):
            continue
        vals = [num(ws.cell(r, c).value) for c in range(2, 6)]
        if all(v is None for v in vals):
            continue
        trimestres.append(f"{anio}-T{TRIMESTRES[m.group(1)]}")
        for k, v in zip(series, vals):
            series[k].append(r1(v))
    if len(trimestres) < 10:
        raise ValueError(f"{path}: {len(trimestres)} trimestres, esperaba 10+")
    return {"trimestres": trimestres, **series}


# ---------------------------------------------------------------- pobreza

def pobreza(path):
    """Una hoja por año, un bloque por trimestre. Se leen las columnas de
    personas (porcentaje y valores absolutos) para pobreza e indigencia."""
    wb = openpyxl.load_workbook(path, data_only=True)
    hojas = [(anio_de(s), s) for s in wb.sheetnames]
    hojas = [(a, s) for a, s in hojas if a and len(s.strip()) <= 6]
    if not hojas:
        raise ValueError(f"{path}: no hay hojas por año (hay {wb.sheetnames})")

    filas = []
    for anio, sn in sorted(hojas):
        ws = wb[sn]
        # cada bloque de trimestre arranca en una fila que lo nombra
        bloques = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column, 40) + 1):
                m = re.match(r"(1er|2do|3er|4to)\.?\s*trimestre", norm(ws.cell(r, c).value))
                if m:
                    bloques.append((TRIMESTRES[m.group(1)], r, c))
        if not bloques:
            continue
        for q, r_tri, c_tri in bloques:
            # dentro del bloque, ubicar las columnas "Personas / Porcentajes" y
            # "Personas / Valores absolutos"
            c_pct = c_abs = None
            for r in range(r_tri, min(r_tri + 4, ws.max_row) + 1):
                for c in range(c_tri, min(c_tri + 12, ws.max_column) + 1):
                    if norm(ws.cell(r, c).value) == "personas":
                        for rr in range(r + 1, min(r + 3, ws.max_row) + 1):
                            for cc in range(c, min(c + 8, ws.max_column) + 1):
                                e = norm(ws.cell(rr, cc).value)
                                if e.startswith("porcentaje") and c_pct is None:
                                    c_pct = cc
                                if e.startswith("valores absolutos") and c_abs is None:
                                    c_abs = cc
            if c_pct is None:
                continue
            def buscar(pref):
                for r in range(r_tri, ws.max_row + 1):
                    if norm(ws.cell(r, 1).value).startswith(pref):
                        return r
                return None
            r_pob = buscar("en situacion de pobreza")
            r_ind = buscar("en situacion de indigencia")
            if r_pob is None or r_ind is None:
                continue
            filas.append({
                "periodo": f"{anio}-T{q}",
                "pob_pct": r1(num(ws.cell(r_pob, c_pct).value), 2),
                "ind_pct": r1(num(ws.cell(r_ind, c_pct).value), 2),
                "pob_abs": num(ws.cell(r_pob, c_abs).value) if c_abs else None,
                "ind_abs": num(ws.cell(r_ind, c_abs).value) if c_abs else None,
            })

    filas = [f for f in filas if f["pob_pct"] is not None]
    filas.sort(key=lambda f: (int(f["periodo"][:4]), int(f["periodo"][-1])))
    if len(filas) < 8:
        raise ValueError(f"{path}: {len(filas)} trimestres de pobreza, esperaba 8+")
    return {
        "periodos":    [f["periodo"] for f in filas],
        "pob_per_pct": [f["pob_pct"] for f in filas],
        "ind_per_pct": [f["ind_pct"] for f in filas],
        "pob_per_abs": [f["pob_abs"] for f in filas],
        "ind_per_abs": [f["ind_abs"] for f in filas],
    }


# ---------------------------------------------------------------- comex

def comex(path):
    _, ws = abrir(path)
    anios, expo, pct = [], [], []
    for r in range(1, ws.max_row + 1):
        a = anio_de(ws.cell(r, 1).value)
        if a is None:
            continue
        usd = num(ws.cell(r, 2).value)
        if usd is None:
            continue
        anios.append(a)
        expo.append(r1(usd / 1e6, 1))                 # el observatorio grafica millones
        pct.append(r1(num(ws.cell(r, 3).value), 4))
    if len(anios) < 10:
        raise ValueError(f"{path}: {len(anios)} años de comex, esperaba 10+")
    return {"anios": anios, "expo": expo, "pct_pgb": pct}


# ------------------------------------------------- industria / masa salarial

def _serie_mensual_por_rama(path):
    """Layout compartido: col A año (sólo en enero), col B mes, col C total,
    col D+ ramas. Devuelve (periodos, total, ramas{nombre: serie})."""
    _, ws = abrir(path)
    fila_ramas = None
    for r in range(1, 8):
        if norm(ws.cell(r, 3).value) == "total":
            fila_ramas = r
            break
    if fila_ramas is None:
        raise ValueError(f"{path}: no encontré la fila de ramas")

    ramas_col = {}
    for c in range(4, ws.max_column + 1):
        nombre = ws.cell(fila_ramas, c).value
        if nombre and len(str(nombre).strip()) > 2:
            ramas_col[str(nombre).strip()] = c

    periodos, total = [], []
    ramas = {k: [] for k in ramas_col}
    anio = None
    for r in range(fila_ramas + 1, ws.max_row + 1):
        a = anio_de(ws.cell(r, 1).value)
        if a:
            anio = a
        mes = MESES.get(norm(ws.cell(r, 2).value))
        if not (mes and anio):
            continue
        t = num(ws.cell(r, 3).value)
        if t is None:
            continue
        periodos.append(f"{anio}-{mes:02d}")
        total.append(r1(t, 2))
        for nombre, c in ramas_col.items():
            ramas[nombre].append(r1(num(ws.cell(r, c).value), 2))
    if len(periodos) < 24:
        raise ValueError(f"{path}: {len(periodos)} meses, esperaba 24+")
    return periodos, total, ramas


def industria(path, ipcba_bloque=None):
    """Ingresos fabriles. Si se pasa la serie del IPCBA, se deflacta para
    obtener la serie a precios constantes que grafica el observatorio."""
    periodos, total, ramas = _serie_mensual_por_rama(path)

    total_const = total
    if ipcba_bloque and ipcba_bloque.get("meses"):
        # se reconstruye un índice de precios encadenado desde las variaciones
        idx, acum = {}, 100.0
        for etiqueta, v in zip(ipcba_bloque["meses"], ipcba_bloque["var_m"]):
            mm, aa = etiqueta.split("-")
            k = f"20{aa}-{MES_CORTO.index(mm) + 1:02d}"
            acum *= (1 + (v or 0) / 100)
            idx[k] = acum
        if idx:
            base = idx[max(idx)]
            total_const = [r1(t * base / idx[p], 2) if p in idx and idx[p] else None
                           for p, t in zip(periodos, total)]

    ult = len(periodos) - 1
    suma = sum(v[ult] for v in ramas.values() if v[ult] is not None) or 1
    pesos = {n: r1(v[ult] / suma * 100, 2) for n, v in ramas.items()
             if v[ult] is not None}
    return {
        "periodos": periodos,
        "total_const": total_const,
        "pesos": pesos,
        "pesos_periodo": periodos[ult],
    }


def masa_salarial(path):
    periodos, total, _ = _serie_mensual_por_rama(path)
    return {"periodos": periodos, "total": total}


# ---------------------------------------------------------------- locales

def locales_evo(path):
    """Serie de la fila 'Total General' por cuatrimestre."""
    _, ws = abrir(path)
    fila_sub = next((r for r in range(1, 8)
                     if "cuatrimestre" in norm(ws.cell(r, 2).value)), None)
    if fila_sub is None:
        raise ValueError(f"{path}: no encontré la fila de cuatrimestres")
    fila_anio = fila_sub - 1

    periodos, cols, anio = [], [], None
    for c in range(2, ws.max_column + 1):
        a = anio_de(ws.cell(fila_anio, c).value)
        if a:
            anio = a
        m = re.match(r"(1er|2do|3er|4to)\.?\s*cuatrimestre", norm(ws.cell(fila_sub, c).value))
        if m and anio:
            periodos.append(f"{anio}-C{TRIMESTRES[m.group(1)]}")
            cols.append(c)

    fila_total = next((r for r in range(1, ws.max_row + 1)
                       if norm(ws.cell(r, 1).value).startswith("total general")), None)
    if fila_total is None:
        raise ValueError(f"{path}: no encontré la fila 'Total General'")
    tasa = [r1(num(ws.cell(fila_total, c).value), 1) for c in cols]
    if len(tasa) < 4:
        raise ValueError(f"{path}: {len(tasa)} cuatrimestres, esperaba 4+")
    return {"periodos": periodos, "tasa": tasa}


def comunas_locales(path):
    """Última hoja de período: una fila por comuna con relevados, ocupados,
    tasa y variación interanual."""
    wb = openpyxl.load_workbook(path, data_only=True)
    hojas = []
    for sn in wb.sheetnames:
        m = re.match(r"(1er|2do|3er|4to)\.?\s*cuatr\.?\s*de\s*((19|20)\d{2})", norm(sn))
        if m:
            hojas.append((int(m.group(2)), TRIMESTRES[m.group(1)], sn))
    if not hojas:
        raise ValueError(f"{path}: no hay hojas de cuatrimestre (hay {wb.sheetnames})")
    anio, q, sn = max(hojas)
    ws = wb[sn]

    fila_hdr = next((r for r in range(1, 8)
                     if norm(ws.cell(r, 1).value).startswith("comuna")), None)
    if fila_hdr is None:
        raise ValueError(f"{path}/{sn}: no encontré el encabezado 'Comuna'")

    data, total = {}, None
    for r in range(fila_hdr + 1, ws.max_row + 1):
        et = str(ws.cell(r, 1).value or "").strip()
        fila = {
            "relev":  num(ws.cell(r, 2).value),
            "ocup":   num(ws.cell(r, 3).value),
            "tasa":   r1(num(ws.cell(r, 4).value), 3),
            "var_ia": r1(num(ws.cell(r, 6).value), 3),
        }
        if fila["tasa"] is None:
            continue
        if norm(et).startswith("total"):
            total = {"relevados": fila["relev"], "ocupados": fila["ocup"],
                     "tasa_ocup": fila["tasa"], "var_ia": fila["var_ia"]}
        elif re.fullmatch(r"\d{1,2}", et):
            data[et] = fila
    if len(data) != 15:
        raise ValueError(f"{path}/{sn}: {len(data)} comunas, esperaba 15")
    return {"periodo": f"{anio}-C{q}", "data": data, "total": total}


# ---------------------------------------------------------------- PGB

def pgb(path):
    """Variación interanual por trimestre: fila de total y una por sector."""
    _, ws = abrir(path)
    fila_hdr = None
    for r in range(1, 12):
        etiquetas = [norm(ws.cell(r, c).value) for c in range(2, 12)]
        if sum(1 for e in etiquetas if "trimestre" in e or re.match(r"(19|20)\d{2}", e)) >= 3:
            fila_hdr = r
            break
    if fila_hdr is None:
        raise ValueError(f"{path}: no encontré el encabezado de trimestres")

    cols, trimestres, anio = [], [], None
    for c in range(2, ws.max_column + 1):
        for r in (fila_hdr - 1, fila_hdr):
            a = anio_de(ws.cell(r, c).value)
            if a:
                anio = a
        m = re.match(r"(1er|2do|3er|4to)\.?\s*trimestre", norm(ws.cell(fila_hdr, c).value))
        if m and anio:
            cols.append(c)
            trimestres.append(f"{anio}-T{TRIMESTRES[m.group(1)]}")
    if len(cols) < 4:
        raise ValueError(f"{path}: {len(cols)} trimestres de PGB, esperaba 4+")

    total, categorias = None, []
    for r in range(fila_hdr + 1, ws.max_row + 1):
        nombre = str(ws.cell(r, 1).value or "").strip()
        if len(nombre) < 4:
            continue
        n = norm(nombre)
        if n.startswith("fuente") or n.startswith("nota"):
            continue
        vals = [r1(num(ws.cell(r, c).value), 4) for c in cols]
        if all(v is None for v in vals):
            continue
        if "producto geografico bruto" in n or n.startswith("total"):
            total = vals
        else:
            categorias.append({"n": nombre, "v": vals})
    if total is None or not categorias:
        raise ValueError(f"{path}: falta el total o los sectores del PGB")

    ult = len(trimestres) - 1
    sectores = sorted(
        [{"nombre": c["n"], "var_ia": c["v"][ult]} for c in categorias
         if c["v"][ult] is not None],
        key=lambda x: -x["var_ia"])
    return {
        "trimestres": trimestres,
        "total": total,
        "categorias": categorias,
        "ultimo_trim": trimestres[ult],
        "ultimo_var": total[ult],
        "sectores_ultimo": sectores,
    }
