#!/usr/bin/env python3
"""Usa el registro oficial de Áreas de Protección Histórica como nube CP-coordenadas.

La ficha BA Data del recurso declara dirección normalizada, barrio, comuna, código
postal tradicional, CPA, latitud y longitud. Se lee el XLSX en modo streaming para
no cargarlo completo en memoria y luego se reutiliza el experimento territorial v1.
"""
from __future__ import annotations

import io
import json
import unicodedata

import openpyxl
import requests

import probar_geolocalizacion_cp_badata as base

APH_PAGE = (
    "https://data.buenosaires.gob.ar/dataset/areas-proteccion-historica/"
    "resource/juqdkmgo-94-resource-xlsx"
)
APH_URL = APH_PAGE + "/download"


def norm(s):
    x = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return "_".join(x.strip().lower().replace("-", "_").split())


def descargar_aph():
    r = requests.get(
        APH_URL,
        headers={"User-Agent": "CEPOES-validacion-territorial/1.0"},
        timeout=180,
        allow_redirects=True,
    )
    r.raise_for_status()
    raw = r.content
    if len(raw) < 100_000 or not raw.startswith(b"PK"):
        raise RuntimeError(f"APH no parece XLSX válido: bytes={len(raw)}, tipo={r.headers.get('content-type')}")

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    encabezado = next(it, None)
    if not encabezado:
        raise RuntimeError("APH XLSX sin encabezado")
    campos = {norm(c): i for i, c in enumerate(encabezado) if c is not None}

    def idx(*candidatos):
        for c in candidatos:
            if c in campos:
                return campos[c]
        return None

    i_cp = idx("codigo_postal", "cod_postal", "cp")
    i_lon = idx("longitud", "long", "lng", "lon", "x")
    i_lat = idx("latitud", "lat", "y")
    i_barrio = idx("barrio")
    if i_cp is None or i_lon is None or i_lat is None:
        raise RuntimeError(f"APH sin columnas CP/lon/lat. Campos={sorted(campos)}")

    filas = []
    leidas = descartadas = 0
    max_i = max(i for i in (i_cp, i_lon, i_lat, i_barrio) if i is not None)
    for row in it:
        leidas += 1
        if len(row) <= max_i:
            descartadas += 1
            continue
        cp = base.cp4(row[i_cp])
        lon = base.numero(row[i_lon])
        lat = base.numero(row[i_lat])
        if cp is None or lon is None or lat is None:
            descartadas += 1
            continue
        if not (-59.0 < lon < -57.5 and -35.0 < lat < -34.0):
            descartadas += 1
            continue
        filas.append({
            "cp": cp,
            "lon": lon,
            "lat": lat,
            "barrio_publicado": str(row[i_barrio] or "").strip() if i_barrio is not None else "",
        })
    wb.close()

    meta = {
        "dataset": "Registro Acumulado de Áreas de Protección Histórica",
        "pagina": APH_PAGE,
        "url_descarga": APH_URL,
        "url_final": r.url,
        "bytes": len(raw),
        "content_type": r.headers.get("content-type", ""),
        "hoja": ws.title,
        "campos_originales": [str(x) if x is not None else "" for x in encabezado],
        "columnas_usadas": {
            "cp": encabezado[i_cp],
            "lon": encabezado[i_lon],
            "lat": encabezado[i_lat],
            "barrio": encabezado[i_barrio] if i_barrio is not None else None,
        },
        "filas_leidas": leidas,
        "filas_validas_cp_coord": len(filas),
        "filas_descartadas": descartadas,
        "cp_distintos": len({x["cp"] for x in filas}),
    }
    if len(filas) < 1000 or meta["cp_distintos"] < 30:
        raise RuntimeError(f"Cobertura APH insuficiente: {meta}")
    return filas, meta


base.descargar_y_leer = descargar_aph
base.BADATA_PAGE = APH_PAGE
base.BADATA_URL = APH_URL
rc = base.main()

out = json.load(open(base.OUTPUT, encoding="utf-8"))
meta = out["fuentes"]["badata_mobiliario_urbano"]
# El nombre de la clave se conserva por compatibilidad con v1, pero el metadato
# identifica inequívocamente que el recurso real es APH.
if meta.get("dataset") != "Registro Acumulado de Áreas de Protección Histórica":
    raise SystemExit("No se utilizó el recurso APH esperado")
if meta.get("cp_distintos", 0) < 30:
    raise SystemExit("APH no aporta cobertura postal suficiente")
raise SystemExit(rc)
