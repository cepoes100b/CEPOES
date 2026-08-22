#!/usr/bin/env python3
"""Prueba CP -> punto representativo -> barrio con fuentes públicas BA Data.

No lee microdatos BCRA/ARCA. Consume el agregado por CP ya producido por la corrida
integral y usa el XLSX vigente de Mobiliario Urbano como nube pública de puntos con
longitud, latitud y código postal. Para cada CP calcula cuatro puntos representativos
(media, mediana, media recortada 10% y punto observado más cercano a la mediana) y
hace point-in-polygon contra el GeoJSON oficial de los 48 barrios.

Los 48 agregados públicos de Mapa de la Deuda son sólo benchmark de QA; nunca se
usan para decidir a qué barrio asignar un CP.
"""
from __future__ import annotations

import io
import json
import math
import statistics
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import requests
from openpyxl import load_workbook
from shapely.geometry import Point, shape

INPUT = Path("diagnostico_universo_territorial_integral.json")
OUTPUT = Path("diagnostico_punto_representativo_cp_mobiliario.json")
MOB_PAGE = "https://data.buenosaires.gob.ar/dataset/mobiliario-urbano/resource/juqdkmgo-1441-resource-xlsx"
MOB_URL = MOB_PAGE + "/download"
BARRIOS_URL = "https://cdn.buenosaires.gob.ar/datosabiertos/datasets/innovacion-transformacion-digital/barrios/barrios.geojson"
LOOKUP_URL = "https://datos.mapadeladeuda.ar/geo/lookup.json"
SLICE_URL = "https://datos.mapadeladeuda.ar/periods/2026-06/slices/barrio_caba/02/default.json"
UA = {"User-Agent": "CEPOES-validacion-territorial/2.1"}


def get(url: str, timeout: int = 120):
    r = requests.get(url, headers=UA, timeout=timeout, allow_redirects=True)
    r.raise_for_status()
    return r


def norm(v) -> str:
    x = unicodedata.normalize("NFKD", str(v or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(x.upper().replace("-", " ").replace(".", " ").split())


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def cp4(v):
    n = num(v)
    if n is None:
        return None
    i = int(round(n))
    return i if 1000 <= i <= 1499 else None


def cargar_mobiliario():
    r = get(MOB_URL, 180)
    raw = r.content
    if len(raw) < 1000 or not raw.startswith(b"PK"):
        raise RuntimeError(f"Mobiliario XLSX inválido: {len(raw)} bytes")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    fields = {norm(x).replace(" ", "_").lower(): i for i, x in enumerate(header) if x is not None}

    def idx(*names):
        for n in names:
            if n in fields:
                return fields[n]
        return None

    ilon = idx("long", "lon", "longitud", "longitude")
    ilat = idx("lat", "latitud", "latitude")
    icp = idx("codigo_postal", "cod_postal", "cp")
    ib = idx("barrio")
    icpa = idx("codigo_postal_argentino", "cpa")
    if ilon is None or ilat is None or icp is None:
        raise RuntimeError(f"Mobiliario XLSX sin lon/lat/CP. Campos={sorted(fields)}")

    puntos = defaultdict(list)
    barrios_fuente = defaultdict(Counter)
    filas = validas = fuera_bbox = 0
    xmin, xmax, ymin, ymax = -58.56, -58.32, -34.72, -34.51
    for row in it:
        filas += 1
        cp = cp4(row[icp] if icp < len(row) else None)
        lon = num(row[ilon] if ilon < len(row) else None)
        lat = num(row[ilat] if ilat < len(row) else None)
        if cp is None or lon is None or lat is None:
            continue
        if not (xmin <= lon <= xmax and ymin <= lat <= ymax):
            fuera_bbox += 1
            continue
        puntos[cp].append((lon, lat))
        if ib is not None and ib < len(row) and row[ib]:
            barrios_fuente[cp][norm(row[ib])] += 1
        validas += 1
    wb.close()

    meta = {
        "pagina": MOB_PAGE,
        "url_descarga": MOB_URL,
        "url_final": r.url,
        "bytes": len(raw),
        "formato": "xlsx",
        "campos": [str(x) if x is not None else None for x in header],
        "columnas": {
            "lon": header[ilon], "lat": header[ilat], "cp": header[icp],
            "barrio": header[ib] if ib is not None else None,
            "cpa": header[icpa] if icpa is not None else None,
        },
        "filas_leidas": filas,
        "filas_cp_coord_validas": validas,
        "filas_coord_fuera_bbox_caba": fuera_bbox,
        "cp_distintos": len(puntos),
    }
    if validas < 100 or len(puntos) < 20:
        raise RuntimeError(f"Cobertura de Mobiliario insuficiente: {meta}")
    return puntos, barrios_fuente, meta


def cargar_barrios_oficiales():
    r = get(BARRIOS_URL, 120)
    obj = r.json()
    polygons = []
    for f in obj.get("features", []):
        props = f.get("properties") or {}
        nombre = props.get("nombre") or props.get("BARRIO") or props.get("barrio")
        if nombre and f.get("geometry"):
            polygons.append((norm(nombre), shape(f["geometry"])))
    if len(polygons) != 48:
        raise RuntimeError(f"GeoJSON oficial: se esperaban 48 barrios y hay {len(polygons)}")
    return polygons, {"url": BARRIOS_URL, "url_final": r.url, "barrios": 48, "bytes": len(r.content)}


def collect_lookup_barrios(obj):
    found = {}
    def visit(node, depth=0):
        if depth > 8:
            return
        if isinstance(node, dict):
            level = str(node.get("level") or node.get("nivel") or "")
            scope = str(node.get("scope") or node.get("scope_id") or "")
            gid = node.get("geo_id") or node.get("id")
            nombre = node.get("nombre") or node.get("name")
            if level == "barrio_caba" and scope in ("", "02") and gid is not None and nombre:
                found[norm(nombre)] = str(gid)
            for child in node.values():
                if isinstance(child, (dict, list)):
                    visit(child, depth + 1)
        elif isinstance(node, list):
            for child in node:
                if isinstance(child, (dict, list)):
                    visit(child, depth + 1)
    visit(obj)
    return found


def cargar_benchmark():
    names = collect_lookup_barrios(get(LOOKUP_URL).json())
    if len(names) != 48:
        raise RuntimeError(f"Lookup Mapa: {len(names)} barrios")
    sl = get(SLICE_URL).json()
    if sl.get("contract") != "mobile-slices-v2":
        raise RuntimeError(f"Contrato Mapa inesperado: {sl.get('contract')}")
    cols = sl["columns"]; aliases = sl.get("aliases", {})
    bench = {}
    for raw in sl["rows"]:
        d = raw if isinstance(raw, dict) else dict(zip(cols, raw))
        gid = str(d["geo_id"])
        bench[gid] = {aliases.get(k, k): v for k, v in d.items() if k != "geo_id"}
    if len(bench) != 48:
        raise RuntimeError(f"Benchmark Mapa: {len(bench)} barrios")
    return names, bench


def trimmed_mean(values, frac=0.10):
    xs = sorted(values); n = len(xs); k = int(n * frac)
    return statistics.fmean(xs[k:n-k]) if n - 2*k >= 1 else statistics.fmean(xs)


def representantes(points):
    lons = [p[0] for p in points]; lats = [p[1] for p in points]
    med = (statistics.median(lons), statistics.median(lats))
    mean = (statistics.fmean(lons), statistics.fmean(lats))
    trim = (trimmed_mean(lons), trimmed_mean(lats))
    medoid = min(points, key=lambda p: (p[0]-med[0])**2 + (p[1]-med[1])**2)
    return {"media": mean, "mediana": med, "media_recortada_10": trim, "punto_cercano_mediana": medoid}


def barrio_de_punto(pt, polygons):
    p = Point(pt[0], pt[1])
    hits = [name for name, poly in polygons if poly.covers(p)]
    return sorted(hits)[0] if hits else None


def pearson(a, b):
    if len(a) < 2:
        return None
    ma = statistics.fmean(a); mb = statistics.fmean(b)
    num = sum((x-ma)*(y-mb) for x, y in zip(a, b))
    da = math.sqrt(sum((x-ma)**2 for x in a)); db = math.sqrt(sum((y-mb)**2 for y in b))
    return num/(da*db) if da and db else None


def compare(agg, bench):
    pairs = {
        "deudores": ("deudores", "deudores_unicos_total", 1),
        "mora": ("personas_mora", "deudores_unicos_mora", 1),
        "deuda": ("deuda_total_pesos", "monto_total", 1/1000),
        "deuda_mora": ("deuda_mora_pesos", "monto_mora", 1/1000),
    }
    gids = sorted(bench); out = {}
    for label, (of, bf, scale) in pairs.items():
        xs = [float(agg.get(g, {}).get(of, 0) or 0)*scale for g in gids]
        ys = [float(bench[g].get(bf, 0) or 0) for g in gids]
        sx = sum(xs); sy = sum(ys); fac = sy/sx if sx else 0
        raw = sum(abs(x-y) for x, y in zip(xs, ys))/sy*100 if sy else None
        nw = sum(abs(x*fac-y) for x, y in zip(xs, ys))/sy*100 if sy else None
        cor = pearson(xs, ys)
        out[label] = {
            "total_asignado": round(sx, 3), "benchmark_total": round(sy, 3),
            "wape_raw_pct": round(raw, 3) if raw is not None else None,
            "wape_distribucion_normalizada_pct": round(nw, 3) if nw is not None else None,
            "correlacion_pearson": round(cor, 4) if cor is not None else None,
        }
    return out


def main():
    src = json.loads(INPUT.read_text(encoding="utf-8"))
    cp_rows = {int(r["clave"]): r for r in src["agregado_cp_1000_1499"]["filas"]}
    puntos, barrios_fuente, meta_mob = cargar_mobiliario()
    polygons, meta_geo = cargar_barrios_oficiales()
    mapa_names, bench = cargar_benchmark()

    oficiales = {name for name, _ in polygons}
    if oficiales != set(mapa_names):
        raise RuntimeError(f"Barrios no reconciliados: BA-no-Mapa={sorted(oficiales-set(mapa_names))}; Mapa-no-BA={sorted(set(mapa_names)-oficiales)}")

    cp_ambiguos = sum(1 for c in barrios_fuente.values() if len(c) > 1)
    mapping = defaultdict(dict); detalle = []
    for cp, pts in puntos.items():
        reps = representantes(pts); row = {"cp": cp, "puntos": len(pts), "barrios_fuente_distintos": len(barrios_fuente.get(cp, {})), "representantes": {}}
        for method, pt in reps.items():
            bn = barrio_de_punto(pt, polygons); gid = mapa_names.get(bn) if bn else None
            if gid:
                mapping[method][cp] = gid
            row["representantes"][method] = {"lon": round(pt[0], 7), "lat": round(pt[1], 7), "barrio": bn, "geo_id": gid}
        detalle.append(row)

    total_deudores = sum(float(r.get("deudores", 0) or 0) for r in cp_rows.values())
    results = {}
    for method, mp in mapping.items():
        agg = defaultdict(lambda: defaultdict(float)); covered = []
        for cp, r in cp_rows.items():
            gid = mp.get(cp)
            if not gid:
                continue
            covered.append(cp)
            for f in ("deudores", "personas_mora", "deuda_total_pesos", "deuda_mora_pesos", "registros"):
                agg[gid][f] += float(r.get(f, 0) or 0)
        cov = sum(float(cp_rows[c].get("deudores", 0) or 0) for c in covered)
        results[method] = {
            "cp_asignados": len(covered),
            "cp_sin_punto_o_fuera_poligono": sorted(set(cp_rows)-set(covered)),
            "cobertura_deudores_pct": round(cov/total_deudores*100, 4) if total_deudores else 0,
            "barrios_con_datos": len(agg),
            "comparacion_48": compare(dict(agg), bench),
        }

    out = {
        "schema": "cepoes-punto-representativo-cp-mobiliario-v2",
        "periodo_benchmark": "2026-06",
        "fuente_mobiliario": meta_mob,
        "fuente_barrios": meta_geo,
        "benchmark": {"slice": SLICE_URL, "lookup": LOOKUP_URL, "uso": "QA solamente"},
        "controles": {"cp_bcra": len(cp_rows), "cp_mobiliario": len(puntos), "cp_mobiliario_multibarrio": cp_ambiguos},
        "resultados": results,
        "representantes_cp_agregados": sorted(detalle, key=lambda x: x["cp"]),
        "privacidad": {"microdatos_bcra_arca_leidos": False, "identificadores_personales_leidos": False, "filas_mobiliario_persistidas": False, "solo_agregados_cp_y_fuentes_publicas": True},
        "advertencia": "Mobiliario Urbano no es un padrón de domicilios. Se usa sólo como hipótesis de punto representativo por CP; su aceptación depende de cobertura y validación independiente contra los 48 barrios.",
    }
    OUTPUT.write_text(json.dumps(out, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
    print(json.dumps({"fuente": meta_mob, "controles": out["controles"], "resultados": results}, ensure_ascii=False, indent=2))
    if max((v["cp_asignados"] for v in results.values()), default=0) < 20:
        raise SystemExit("Cobertura insuficiente para evaluar la hipótesis")


if __name__ == "__main__":
    main()
